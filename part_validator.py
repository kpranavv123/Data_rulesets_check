"""
Part Table Ruleset Validator
-----------------------------
Input  : .tab file (tab-delimited)
Output : Excel file with error-highlighted cells and an "Errors" summary column

Rulesets implemented (from Part (FG) sheet):
  P_RS_1  Name           – Not blank; FERT → 14xxxxxxxxxxxxxx, HAWA → 15xxxxxxxxxxxxxx
  P_RS_2  Site           – Not blank
  P_RS_3  Description    – Not blank
  P_RS_4  PartClass      – Not blank; must be FERT or HAWA
  P_RS_5  ProductFamily  – Not blank
  P_RS_6  UnitOfMeasure  – Not blank; must be in {KG, CV, TO, EA, PAL}
  P_RS_7  TCPL_MRPTYPE   – Not blank; must be ND or PD
  P_RS_8  ProcurementType– Not blank
  P_RS_9  ABCCode        – Not blank
  P_RS_10 IBPSTATUS      – Must be IBP or blank
  P_RS_11 XPLANTMATSTATUS– Must be 2 or blank
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
import os

# ── Colours ──────────────────────────────────────────────────────────────────
ERROR_FILL   = PatternFill("solid", start_color="FF4C4C", end_color="FF4C4C")   # red
HEADER_FILL  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")   # dark blue
META_FILL    = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")   # mid blue
ERR_COL_FILL = PatternFill("solid", start_color="FFD966", end_color="FFD966")   # amber header
VALID_FILL   = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")   # light green

HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
META_FONT    = Font(name="Arial", bold=True, color="FFFFFF", size=9)
DATA_FONT    = Font(name="Arial", size=9)
ERR_FONT     = Font(name="Arial", bold=True, size=9, color="7B0000")

THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

VALID_UOMS    = {"KG", "CV", "TO", "EA", "PAL"}
VALID_MRP     = {"ND", "PD"}
VALID_CLASS   = {"FERT", "HAWA"}
VALID_IBP     = {"IBP", ""}
VALID_XPLANT  = {"2", ""}

FERT_MIN, FERT_MAX = 14_000_000_000_000, 14_999_999_999_999
HAWA_MIN, HAWA_MAX = 15_000_000_000_000, 15_999_999_999_999


# ── Helpers ───────────────────────────────────────────────────────────────────
def blank(val) -> bool:
    return pd.isna(val) or str(val).strip() == ""


def sval(val) -> str:
    return "" if pd.isna(val) else str(val).strip()


# ── Rule Functions ─────────────────────────────────────────────────────────────
def check_name(row) -> str | None:
    v = sval(row.get("Name", ""))
    part_class = sval(row.get("PartClass", "")).upper()
    if not v:
        return "Name is blank"
    try:
        num = int(v)
    except ValueError:
        return "Name must be numeric"
    if part_class == "FERT" and not (FERT_MIN <= num <= FERT_MAX):
        return f"Name out of FERT range (14000000000000–14999999999999)"
    if part_class == "HAWA" and not (HAWA_MIN <= num <= HAWA_MAX):
        return f"Name out of HAWA range (15000000000000–15999999999999)"
    return None


def check_not_blank(field_name):
    def _check(row) -> str | None:
        if blank(row.get(field_name, "")):
            return f"{field_name} is blank"
        return None
    return _check


def check_part_class(row) -> str | None:
    v = sval(row.get("PartClass", "")).upper()
    if not v:
        return "PartClass is blank"
    if v not in VALID_CLASS:
        return f"PartClass must be FERT or HAWA (got '{v}')"
    return None


def check_uom(row) -> str | None:
    v = sval(row.get("UnitOfMeasure", "")).upper()
    if not v:
        return "UnitOfMeasure is blank"
    if v not in VALID_UOMS:
        return f"UnitOfMeasure must be one of {sorted(VALID_UOMS)} (got '{v}')"
    return None


def check_mrptype(row) -> str | None:
    v = sval(row.get("TCPL_MRPTYPE", "")).upper()
    if not v:
        return "TCPL_MRPTYPE is blank"
    if v not in VALID_MRP:
        return f"TCPL_MRPTYPE must be ND or PD (got '{v}')"
    return None


def check_ibpstatus(row) -> str | None:
    v = sval(row.get("IBPSTATUS", "")).upper()
    if v not in VALID_IBP:
        return f"IBPSTATUS must be 'IBP' or blank (got '{v}')"
    return None


def check_xplant(row) -> str | None:
    v = sval(row.get("XPLANTMATSTATUS", ""))
    if v not in VALID_XPLANT:
        return f"XPLANTMATSTATUS must be '2' or blank (got '{v}')"
    return None


# Map: column name → (ruleset_id, checker function, display label)
COLUMN_RULES = {
    "Name":             ("P_RS_1",  check_name,                        "Name"),
    "Site":             ("P_RS_2",  check_not_blank("Site"),            "Site"),
    "Description":      ("P_RS_3",  check_not_blank("Description"),     "Description"),
    "PartClass":        ("P_RS_4",  check_part_class,                   "PartClass"),
    "ProductFamily":    ("P_RS_5",  check_not_blank("ProductFamily"),   "ProductFamily"),
    "UnitOfMeasure":    ("P_RS_6",  check_uom,                          "UnitOfMeasure"),
    "TCPL_MRPTYPE":     ("P_RS_7",  check_mrptype,                      "TCPL_MRPTYPE"),
    "ProcurementType":  ("P_RS_8",  check_not_blank("ProcurementType"), "ProcurementType"),
    "ABCCode":          ("P_RS_9",  check_not_blank("ABCCode"),         "ABCCode"),
    "IBPSTATUS":        ("P_RS_10", check_ibpstatus,                    "IBPSTATUS"),
    "XPLANTMATSTATUS":  ("P_RS_11", check_xplant,                       "XPLANTMATSTATUS"),
}


# ── Main Processor ────────────────────────────────────────────────────────────
def process(tab_path: str, output_path: str):
    df = pd.read_csv(tab_path, sep="\t", dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]

    # ── Validate each row ────────────────────────────────────────────────────
    # error_map[row_idx][col_name] = error message
    error_map: dict[int, dict[str, str]] = {}
    error_summary: list[str] = []          # per-row list of field names with errors

    for i, row in df.iterrows():
        row_errors: dict[str, str] = {}
        for col, (rs_id, checker, label) in COLUMN_RULES.items():
            if col in df.columns:
                msg = checker(row)
                if msg:
                    row_errors[col] = msg
        error_map[i] = row_errors
        if row_errors:
            error_summary.append(", ".join(row_errors.keys()))
        else:
            error_summary.append("")

    df["Validation_Errors"] = error_summary

    # ── Build Excel workbook ─────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Part (FG) – Validated"

    # ─ Meta info rows ────────────────────────────────────────────────────────
    meta_lines = [
        ("Interface :", "Part (FG)"),
        ("Description :", "The Part table identifies a unique part and site. "
                          "A record includes: name, description, price, and cost data."),
        ("Source File :", os.path.basename(tab_path)),
        ("Total Records :", len(df)),
        ("Records with Errors :", sum(1 for e in error_summary if e)),
    ]
    for line in meta_lines:
        ws.append(list(line))
    ws.append([])  # blank separator

    meta_row_count = len(meta_lines) + 1   # +1 for blank row

    # ─ Column headers ────────────────────────────────────────────────────────
    all_cols  = list(df.columns)
    data_cols = [c for c in all_cols if c != "Validation_Errors"]
    ordered   = data_cols + ["Validation_Errors"]

    header_row_num = meta_row_count + 1
    ws.append(ordered)

    # ─ Style meta rows ───────────────────────────────────────────────────────
    for r in range(1, meta_row_count):
        for cell in ws[r]:
            cell.fill      = META_FILL
            cell.font      = META_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border    = BORDER

    # ─ Style header row ──────────────────────────────────────────────────────
    for ci, col_name in enumerate(ordered, start=1):
        cell = ws.cell(row=header_row_num, column=ci)
        cell.value     = col_name
        cell.fill      = HEADER_FILL if col_name != "Validation_Errors" else ERR_COL_FILL
        cell.font      = HEADER_FONT if col_name != "Validation_Errors" else Font(name="Arial", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER

    # ─ Data rows ─────────────────────────────────────────────────────────────
    for df_idx, (_, row) in enumerate(df.iterrows()):
        excel_row = header_row_num + 1 + df_idx
        row_has_error = bool(error_map[df_idx])

        for ci, col_name in enumerate(ordered, start=1):
            cell = ws.cell(row=excel_row, column=ci)
            cell.value     = sval(row[col_name])
            cell.font      = DATA_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border    = BORDER

            if col_name == "Validation_Errors":
                if row_has_error:
                    cell.fill = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
                    cell.font = ERR_FONT
                else:
                    cell.fill = VALID_FILL
            elif col_name in error_map[df_idx]:
                # Highlight the specific errored cell
                cell.fill = ERROR_FILL
                cell.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
                # Attach a comment with the error message
                from openpyxl.comments import Comment
                err_msg  = error_map[df_idx][col_name]
                rs_id    = COLUMN_RULES[col_name][0]
                comment  = Comment(f"[{rs_id}] {err_msg}", "Validator")
                comment.width  = 250
                comment.height = 60
                cell.comment   = comment
            else:
                cell.fill = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")

    # ─ Column widths ─────────────────────────────────────────────────────────
    col_widths = {
        "Name": 22, "Site": 10, "Description": 38, "PartClass": 14,
        "ProductFamily": 18, "UnitOfMeasure": 16, "TCPL_MRPTYPE": 16,
        "ProcurementType": 18, "ABCCode": 12, "IBPSTATUS": 14,
        "XPLANTMATSTATUS": 20, "Validation_Errors": 45,
    }
    for ci, col_name in enumerate(ordered, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(col_name, 16)

    ws.row_dimensions[header_row_num].height = 28

    # ─ Freeze pane below header ──────────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=header_row_num + 1, column=1)

    # ── Legend sheet ─────────────────────────────────────────────────────────
    ls = wb.create_sheet("Ruleset Legend")
    legend_headers = ["Ruleset No.", "Field Name", "Rule Description"]
    legend_data = [
        ("P_RS_1",  "Name",             "Not blank. FERT → 14000000000000–14999999999999; HAWA → 15000000000000–15999999999999"),
        ("P_RS_2",  "Site",             "Not blank"),
        ("P_RS_3",  "Description",      "Not blank"),
        ("P_RS_4",  "PartClass",        "Not blank; must be FERT or HAWA"),
        ("P_RS_5",  "ProductFamily",    "Not blank"),
        ("P_RS_6",  "UnitOfMeasure",    "Not blank; must be one of: KG, CV, TO, EA, PAL"),
        ("P_RS_7",  "TCPL_MRPTYPE",     "Not blank; must be ND or PD"),
        ("P_RS_8",  "ProcurementType",  "Not blank"),
        ("P_RS_9",  "ABCCode",          "Not blank"),
        ("P_RS_10", "IBPSTATUS",        "Must be 'IBP' or blank"),
        ("P_RS_11", "XPLANTMATSTATUS",  "Must be '2' or blank"),
    ]
    ls.append(legend_headers)
    for ci, h in enumerate(legend_headers, 1):
        cell = ls.cell(row=1, column=ci)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    for row_data in legend_data:
        ls.append(list(row_data))

    for ri in range(2, len(legend_data) + 2):
        for ci in range(1, 4):
            cell = ls.cell(row=ri, column=ci)
            cell.font      = DATA_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border    = BORDER

    ls.column_dimensions["A"].width = 14
    ls.column_dimensions["B"].width = 22
    ls.column_dimensions["C"].width = 70
    ls.row_dimensions[1].height = 22

    # ── Save ─────────────────────────────────────────────────────────────────
    wb.save(output_path)

    # ── Console summary ──────────────────────────────────────────────────────
    total   = len(df)
    errored = sum(1 for e in error_summary if e)
    print(f"\n{'='*55}")
    print(f"  Part Ruleset Validation Complete")
    print(f"{'='*55}")
    print(f"  Total records   : {total}")
    print(f"  Records OK      : {total - errored}")
    print(f"  Records w/ errors: {errored}")
    print(f"  Output saved to : {output_path}")
    print(f"{'='*55}\n")

    print("  Row-level error summary:")
    for i, err in enumerate(error_summary, start=1):
        status = f"  Row {i:>3}: {'✓ OK' if not err else '✗ ' + err}"
        print(status)


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    tab_file = sys.argv[1] if len(sys.argv) > 1 else "part_sample.tab"
    out_file = sys.argv[2] if len(sys.argv) > 2 else "part_validated.xlsx"
    process(tab_file, out_file)
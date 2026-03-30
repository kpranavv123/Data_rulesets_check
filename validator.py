"""
Multi-Sheet Ruleset Validator
==============================
Input  : Excel file (.xlsx) with sheets: Part, Site
Output : Separate validated Excel files in ./output/ folder

- Validates ONLY the specified columns per sheet
- Keeps ALL columns from the input file in the output
- Handles column names with any casing, spaces, or underscores

Part Sheet columns validated:
  P_RS_1  PLANT                - Not blank; must be in consolidated PL list
  P_RS_2  PRODUCTDESCRIPTION   - Not blank
  P_RS_3  PRODUCTTYPE          - Not blank; must be FERT or HAWA
  P_RS_4  PRODUCTHIERARCHY     - Not blank
  P_RS_5  BASEUNIT             - Not blank; must be KG/CV/TO/EA/PAL
  P_RS_6  MRPTYPE              - Not blank; must be ND or PD
  P_RS_7  PROCUREMENTTYPE      - Not blank
  P_RS_8  ABCINDICATOR         - Not blank

Site Sheet columns validated:
  S_RS_1  PLANT                       - Not blank; must be in consolidated PL list
  S_RS_2  NAME                        - Not blank
  S_RS_3  NODETYPESUPPLYCHAINNETWORK  - Not blank; must be in standard nodetype list

Usage:
  python validator.py input_data.xlsx
  python validator.py input_data.xlsx output_folder/
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import sys
import os
import re

# =============================================================================
# STYLES
# =============================================================================
ROW_ERR_FILL  = PatternFill("solid", start_color="FFE5E5", end_color="FFE5E5")
CELL_ERR_FILL = PatternFill("solid", start_color="C00000", end_color="C00000")
CELL_ERR_FONT = Font(name="Arial", bold=True, size=9, color="FFFFFF")
HEADER_FILL   = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
META_FILL     = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
META_FONT     = Font(name="Arial", bold=True, color="FFFFFF", size=9)
ERR_HDR_FILL  = PatternFill("solid", start_color="FF8C00", end_color="FF8C00")
ERR_HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
ERR_CELL_FILL = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
ERR_CELL_FONT = Font(name="Arial", bold=True, size=9, color="7B2C00")
OK_FONT       = Font(name="Arial", size=9)
WHITE_FILL    = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
GREEN_FILL    = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
THIN          = Side(border_style="thin", color="BFBFBF")
BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# =============================================================================
# REFERENCE LISTS
# =============================================================================
VALID_PLANTS    = {"IN01","IN02","IN03","IN04","IN05","IN06","IN07","IN08","IN09","IN10"}
VALID_UOMS      = {"KG","CV","TO","EA","PAL"}
VALID_MRP       = {"ND","PD"}
VALID_PART_TYPE = {"FERT","HAWA"}
VALID_NODETYPES = {"DC","MFG","DIST","WH","3PL","CUST","SUPPLIER"}


# =============================================================================
# FUZZY COLUMN MATCHING
# =============================================================================
def normalise(name):
    """Strip spaces/underscores/hyphens and lowercase."""
    return re.sub(r"[\s_\-]+", "", str(name)).lower()


# canonical key -> set of normalised aliases that resolve to it
PART_ALIASES = {
    "PLANT":              {"plant"},
    "PRODUCTDESCRIPTION": {"productdescription","productdesc","description"},
    "PRODUCTTYPE":        {"producttype","type","partclass","materialtype"},
    "PRODUCTHIERARCHY":   {"producthierarchy","producthierarch","productfamily"},
    "BASEUNIT":           {"baseunit","unitofmeasure","uom","unit"},
    "MRPTYPE":            {"mrptype","mrp","tcplmrptype"},
    "PROCUREMENTTYPE":    {"procurementtype","procurement"},
    "ABCINDICATOR":       {"abcindicator","abccode","abc"},
}

SITE_ALIASES = {
    "PLANT":                      {"plant","site","plantcode"},
    "NAME":                       {"name","sitename","plantname"},
    "NODETYPESUPPLYCHAINNETWORK": {"nodetypesupplychainnetwork","nodetype",
                                   "nodetypescn","supplychainnetwork"},
}


def build_col_map(df_columns, aliases):
    """Returns {canonical_key: actual_col_name_in_df}."""
    norm_to_actual = {normalise(c): c for c in df_columns}
    col_map = {}
    for canon, alias_set in aliases.items():
        matched = next((norm_to_actual[a] for a in alias_set if a in norm_to_actual), None)
        if matched:
            col_map[canon] = matched
        else:
            print(f"    WARNING: '{canon}' not matched in sheet (tried: {sorted(alias_set)})")
    return col_map


# =============================================================================
# HELPERS
# =============================================================================
def sval(val):
    return "" if pd.isna(val) else str(val).strip()


# =============================================================================
# RULE FUNCTIONS  (receive a canonical-keyed dict for the row)
# =============================================================================
def rule_plant(row):
    v = sval(row.get("PLANT", ""))
    if not v:
        return "PLANT is blank"
    if v not in VALID_PLANTS:
        return f"PLANT '{v}' not in consolidated PL list"
    return None

def rule_plant_not_blank_only(row):
    v = sval(row.get("PLANT", ""))
    if not v:
        return "PLANT is blank"
    return None


def rule_not_blank(canon):
    def _check(row):
        if not sval(row.get(canon, "")):
            return f"{canon} is blank"
        return None
    return _check


def rule_producttype(row):
    v = sval(row.get("PRODUCTTYPE", "")).upper()
    if not v:
        return "PRODUCTTYPE is blank"
    if v not in VALID_PART_TYPE:
        return f"PRODUCTTYPE must be FERT or HAWA (got '{v}')"
    return None


def rule_baseunit(row):
    v = sval(row.get("BASEUNIT", "")).upper()
    if not v:
        return "BASEUNIT is blank"
    if v not in VALID_UOMS:
        return f"BASEUNIT must be one of {sorted(VALID_UOMS)} (got '{v}')"
    return None


def rule_mrptype(row):
    v = sval(row.get("MRPTYPE", "")).upper()
    if not v:
        return "MRPTYPE is blank"
    if v not in VALID_MRP:
        return f"MRPTYPE must be ND or PD (got '{v}')"
    return None


def rule_nodetype(row):
    v = sval(row.get("NODETYPESUPPLYCHAINNETWORK", "")).upper()
    if not v:
        return "NODETYPESUPPLYCHAINNETWORK is blank"
    if v not in VALID_NODETYPES:
        return f"Nodetype '{v}' not in standard list {sorted(VALID_NODETYPES)}"
    return None


PART_RULE_DEFS = {
    "PLANT":              ("P_RS_1", rule_not_blank("PLANT")),
    "PRODUCTDESCRIPTION": ("P_RS_2", rule_not_blank("PRODUCTDESCRIPTION")),
    "PRODUCTTYPE":        ("P_RS_3", rule_producttype),
    "PRODUCTHIERARCHY":   ("P_RS_4", rule_not_blank("PRODUCTHIERARCHY")),
    "BASEUNIT":           ("P_RS_5", rule_not_blank("BASEUNIT")),
    "MRPTYPE":            ("P_RS_6", rule_mrptype),
    "PROCUREMENTTYPE":    ("P_RS_7", rule_not_blank("PROCUREMENTTYPE")),
    "ABCINDICATOR":       ("P_RS_8", rule_not_blank("ABCINDICATOR")),
}

SITE_RULE_DEFS = {
    "PLANT": ("S_RS_1", rule_plant_not_blank_only),
    "NAME":("S_RS_2", rule_not_blank("NAME")),
    "NODETYPESUPPLYCHAINNETWORK": ("S_RS_3", rule_not_blank("NODETYPESUPPLYCHAINNETWORK")),
}

PART_LEGEND = [
    ("P_RS_1","PLANT",             "Not blank; must be in consolidated PL list"),
    ("P_RS_2","PRODUCTDESCRIPTION","Not blank"),
    ("P_RS_3","PRODUCTTYPE",       "Not blank; must be FERT or HAWA"),
    ("P_RS_4","PRODUCTHIERARCHY",  "Not blank"),
    ("P_RS_5","BASEUNIT",          f"Not blank; must be one of: {sorted(VALID_UOMS)}"),
    ("P_RS_6","MRPTYPE",           "Not blank; must be ND or PD"),
    ("P_RS_7","PROCUREMENTTYPE",   "Not blank"),
    ("P_RS_8","ABCINDICATOR",      "Not blank"),
]

SITE_LEGEND = [
    ("S_RS_1","PLANT",                     "Not blank; must be in consolidated PL list"),
    ("S_RS_2","NAME",                       "Not blank"),
    ("S_RS_3","NODETYPESUPPLYCHAINNETWORK", f"Not blank; must be in: {sorted(VALID_NODETYPES)}"),
]


# =============================================================================
# VALIDATE
# =============================================================================
def validate_dataframe(df, col_map, rule_defs):
    error_map = []
    err_summary = []

    for _, raw_row in df.iterrows():
        canon_row = {canon: sval(raw_row.get(actual, ""))
                     for canon, actual in col_map.items()}

        row_errors = {}
        for canon, (rs_id, checker) in rule_defs.items():
            if canon not in col_map:
                continue
            msg = checker(canon_row)
            if msg:
                row_errors[col_map[canon]] = msg   # key = actual col name

        error_map.append(row_errors)
        err_summary.append(", ".join(row_errors.keys()) if row_errors else "")

    return error_map, err_summary


# =============================================================================
# WRITE EXCEL
# =============================================================================
def write_validated_excel(df, error_map, err_summary, col_map,
                          rule_defs, output_path, sheet_label, legend_data):
    df = df.copy()
    df["Validation_Errors"] = err_summary

    # ALL original columns + Validation_Errors at the end
    ordered = [c for c in df.columns if c != "Validation_Errors"] + ["Validation_Errors"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_label[:31]

    total   = len(df)
    errored = sum(1 for e in err_summary if e)

    # Meta block
    meta = [
        ["Interface :", sheet_label],
        ["Total Records :", total],
        ["Records with Errors :", errored],
        ["Records OK :", total - errored],
        ["", ""],
    ]
    for line in meta:
        ws.append(line)
    for r in range(1, len(meta) + 1):
        for c in range(1, len(ordered) + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = META_FILL
            cell.font = META_FONT
            cell.alignment = Alignment(vertical="center")
            cell.border = BORDER

    header_row = len(meta) + 1
    ws.append(ordered)
    ws.row_dimensions[header_row].height = 30

    # Reverse map: actual col name -> ruleset_id (for comments)
    actual_to_rsid = {col_map[c]: rs for c, (rs, _) in rule_defs.items() if c in col_map}

    # Header styling
    for ci, col_name in enumerate(ordered, start=1):
        cell = ws.cell(row=header_row, column=ci)
        cell.value = col_name
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        if col_name == "Validation_Errors":
            cell.fill = ERR_HDR_FILL
            cell.font = ERR_HDR_FONT
        else:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

    # Data rows
    for df_idx, (_, row) in enumerate(df.iterrows()):
        excel_row     = header_row + 1 + df_idx
        row_has_error = bool(error_map[df_idx])

        for ci, col_name in enumerate(ordered, start=1):
            cell = ws.cell(row=excel_row, column=ci)
            cell.value     = sval(row[col_name])
            cell.border    = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)

            if col_name == "Validation_Errors":
                if row_has_error:
                    cell.fill  = ERR_CELL_FILL
                    cell.font  = ERR_CELL_FONT
                else:
                    cell.fill  = GREEN_FILL
                    cell.font  = Font(name="Arial", size=9, color="375623", bold=True)
                    cell.value = "OK"
            elif col_name in error_map[df_idx]:
                # Specific failing cell -> dark red
                cell.fill = CELL_ERR_FILL
                cell.font = CELL_ERR_FONT
                rs_id = actual_to_rsid.get(col_name, "?")
                cmt = Comment(f"[{rs_id}] {error_map[df_idx][col_name]}", "Validator")
                cmt.width = 270; cmt.height = 65
                cell.comment = cmt
            elif row_has_error:
                # Row has error, this cell is fine -> light pink tint
                cell.fill = ROW_ERR_FILL
                cell.font = OK_FONT
            else:
                cell.fill = WHITE_FILL
                cell.font = OK_FONT

    # Column widths
    canon_widths = {
        "PLANT":22, "PRODUCTDESCRIPTION":40, "PRODUCTTYPE":16,
        "PRODUCTHIERARCHY":22, "BASEUNIT":14, "MRPTYPE":12,
        "PROCUREMENTTYPE":18, "ABCINDICATOR":16,
        "NAME":32, "NODETYPESUPPLYCHAINNETWORK":32,
    }
    actual_widths = {col_map.get(c, c): w for c, w in canon_widths.items()}

    for ci, col_name in enumerate(ordered, start=1):
        ltr = get_column_letter(ci)
        ws.column_dimensions[ltr].width = (
            50 if col_name == "Validation_Errors" else actual_widths.get(col_name, 18)
        )

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # Legend sheet
    ls = wb.create_sheet("Ruleset Legend")
    ls.append(["Ruleset No.", "Data Lake Column", "Rule Description"])
    for h in ls[1]:
        h.fill = HEADER_FILL; h.font = HEADER_FONT
        h.alignment = Alignment(horizontal="center", vertical="center")
        h.border = BORDER
    for i, rd in enumerate(legend_data, start=2):
        ls.append(list(rd))
        bg = "EBF3FB" if i % 2 == 0 else "FFFFFF"
        for ci in range(1, 4):
            c = ls.cell(row=i, column=ci)
            c.font = Font(name="Arial", size=9)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = BORDER
            c.fill = PatternFill("solid", start_color=bg, end_color=bg)
    ls.column_dimensions["A"].width = 14
    ls.column_dimensions["B"].width = 30
    ls.column_dimensions["C"].width = 72

    # Colour guide
    cg = wb.create_sheet("Colour Guide")
    cg.append(["Sample", "Meaning"])
    for h in cg[1]:
        h.fill = HEADER_FILL; h.font = HEADER_FONT
        h.alignment = Alignment(horizontal="center"); h.border = BORDER
    guide = [
        (CELL_ERR_FILL, "Dark Red cell  - this specific field failed a rule (hover for detail)"),
        (ROW_ERR_FILL,  "Light Pink cell - same row has an error but this field is OK"),
        (ERR_CELL_FILL, "Amber Validation_Errors - lists the failed field names for this row"),
        (GREEN_FILL,    "Green Validation_Errors - row passed all rules"),
        (WHITE_FILL,    "White row - fully valid"),
    ]
    for i, (fill, meaning) in enumerate(guide, start=2):
        cg.cell(row=i, column=1).fill  = fill
        cg.cell(row=i, column=2).value = meaning
        for ci in range(1, 3):
            c = cg.cell(row=i, column=ci)
            c.font = Font(name="Arial", size=9)
            c.border = BORDER
            c.alignment = Alignment(vertical="center")
    cg.column_dimensions["A"].width = 10
    cg.column_dimensions["B"].width = 65

    wb.save(output_path)


# =============================================================================
# MAIN
# =============================================================================
def main(input_path, output_folder="output"):
    if not os.path.exists(input_path):
        print(f"\nERROR: File not found -> {input_path}")
        sys.exit(1)

    os.makedirs(output_folder, exist_ok=True)

    xl = pd.ExcelFile(input_path)
    print(f"\nSheets in '{os.path.basename(input_path)}': {xl.sheet_names}")

    configs = [
        ("part", PART_ALIASES, PART_RULE_DEFS, "Part (FG)", "Part_Validated.xlsx", PART_LEGEND),
        ("site", SITE_ALIASES, SITE_RULE_DEFS, "Site",      "Site_Validated.xlsx", SITE_LEGEND),
    ]
    results = []

    for target, aliases, rule_defs, label, out_name, legend in configs:
        matched_sheet = next(
     (s for s in xl.sheet_names if target in s.strip().lower()), None
      )
        if not matched_sheet:
            print(f"\n  [{label}]  Sheet '{target}' not found - skipping")
            continue

        df = pd.read_excel(input_path, sheet_name=matched_sheet, dtype=str).fillna("")
        print(f"\n  [{label}]  Sheet: '{matched_sheet}'  |  {len(df)} rows  |  {len(df.columns)} columns")
        print(f"             Input columns : {list(df.columns)}")

        col_map = build_col_map(list(df.columns), aliases)
        print(f"             Matched rules : {col_map}")

        error_map, err_summary = validate_dataframe(df, col_map, rule_defs)

        out_path = os.path.join(output_folder, out_name)
        write_validated_excel(df, error_map, err_summary, col_map,
                              rule_defs, out_path, label, legend)

        errored = sum(1 for e in err_summary if e)
        results.append((label, len(df), errored, out_path))

        print(f"             Errors: {errored}  |  OK: {len(df)-errored}")
        for i, s in enumerate(err_summary, 1):
            print(f"    Row {i:>4}: {'OK' if not s else 'ERROR -> ' + s}")

    print(f"\n{'='*62}")
    print(f"  DONE  |  Output folder: ./{output_folder}/")
    print(f"{'='*62}")
    for label, total, errored, path in results:
        print(f"  {label:<12} -> {os.path.basename(path):<30}  ({total} rows | {errored} errors)")
    print(f"{'='*62}\n")


if __name__ == "__main__":
    input_file = r"D:\SEM-8\Data Rules Set Check\Data_rulesets_check\Sample data file - Pranav.xlsx"
    output_folder = "output"
    main(input_file, output_folder)
"""
PRODUCTHIERARCHY (FG) - Excel Data Validation Tool
Validates fields based on material type rules and generates a formatted Excel report.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from typing import Optional
import os


# ─────────────────────────────────────────────
#  FILE PATHS  –  update these
# ─────────────────────────────────────────────
INPUT_FILE  = r"D:\SEM-8\Data Rules Set Check\Data_rulesets_check\Excel_Files\PRODUCTHIERARCHY_FG.xlsx"
OUTPUT_FILE = r"D:\SEM-8\Data Rules Set Check\Data_rulesets_check\Validated_PRODUCTHIERARCHY_FG.xlsx"


# ─────────────────────────────────────────────
#  CONSTANTS & CONFIGURATION
# ─────────────────────────────────────────────

VALID_MATERIAL_TYPES = {"FERT", "HAWA"}
VALID_IBP_STATUSES   = {"IBP", ""}

NOT_BLANK_FIELDS = [
    "MATERIALNUMBER", "MATERIALDESCRIPTION", "PRODUCTGROUP", "MATLGRPDESC",
    "DIVISION", "DIVISIONDESCRIPTION", "PRODUCTTYPE", "PRODUCT_HIERARCHY_KEY",
    "CATEGORY", "CATEGORYDESCRIPTION", "PRODUCT", "PRODUCTDESCRIPTION",
    "VARIANT", "VARIANTDESCRIPTION", "BRAND", "BRANDDESCRIPTION",
    "SUBBRAND", "SUBBRANDDESCRIPTION", "BRANDVARIANT", "BRANDVARIANTDESCRIPTION",
    "PACKSIZE", "PACKSIZEDESCRIPTION", "MARKETSKU", "MARKETSKUDESCRIPTION",
    "SUPPLY_FAMILY",
]

# ─────────────────────────────────────────────
#  Rules content — used for Rule_Set sheet
#  (same merged-cell template as Part & Site)
# ─────────────────────────────────────────────
RULES_CONTENT = {
    **{f: ["Must not be blank for FERT/HAWA material types."] for f in NOT_BLANK_FIELDS},
    "MATERIALTYPE": [
        "Must not be blank.",
        "Value must be either FERT or HAWA.",
    ],
    "IBPSTATUS": [
        "Allowed values: IBP or blank.",
        "Any other value is treated as an error.",
    ],
}

# ─────────────────────────────────────────────
#  Colours  (aligned with Part & Site scripts)
# ─────────────────────────────────────────────
RED_FILL       = PatternFill("solid", start_color="FF0000", end_color="FF0000")
ROW_ERROR_FILL = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
HDR_FILL       = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
RULE_FILL      = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
TITLE_FILL     = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
WHITE_FILL     = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
TOTAL_FILL     = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
STATS_FILL     = PatternFill("solid", start_color="EDEDED", end_color="EDEDED")

HDR_FONT    = Font(bold=True, name="Arial")
BODY_FONT   = Font(name="Arial", size=10)
ERR_FONT    = Font(name="Arial", size=10, bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


# ─────────────────────────────────────────────
#  HELPER UTILITIES
# ─────────────────────────────────────────────

def is_blank(value) -> bool:
    if value is None:
        return True
    if isinstance(value, float):
        import math
        return math.isnan(value)
    return str(value).strip() == ""


def style_header_row(ws, row: int, num_cols: int):
    for col in range(1, num_cols + 1):
        cell           = ws.cell(row=row, column=col)
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER


def auto_width(ws, min_w=10, max_w=50):
    for col in ws.columns:
        length = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)


# ─────────────────────────────────────────────
#  VALIDATION ENGINE
# ─────────────────────────────────────────────

class FieldValidator:
    def __init__(self, field: str, rule_description: str):
        self.field            = field
        self.rule_description = rule_description

    def validate(self, row: pd.Series) -> Optional[str]:
        raise NotImplementedError


class NotBlankValidator(FieldValidator):
    def validate(self, row: pd.Series) -> Optional[str]:
        mat_type = str(row.get("MATERIALTYPE", "")).strip().upper()
        if mat_type not in VALID_MATERIAL_TYPES:
            return None
        value = row.get(self.field)
        if is_blank(value):
            return f"{self.field}: Must not be blank for FERT/HAWA materials"
        return None


class MaterialTypeValidator(FieldValidator):
    def validate(self, row: pd.Series) -> Optional[str]:
        value = str(row.get("MATERIALTYPE", "")).strip().upper()
        if is_blank(value):
            return "MATERIALTYPE: Field must not be blank"
        if value not in VALID_MATERIAL_TYPES:
            return f"MATERIALTYPE: Invalid value '{value}' – must be FERT or HAWA"
        return None


class IBPStatusValidator(FieldValidator):
    def validate(self, row: pd.Series) -> Optional[str]:
        raw   = row.get("IBPSTATUS", "")
        value = "" if is_blank(raw) else str(raw).strip()
        if value not in VALID_IBP_STATUSES and value.upper() not in VALID_IBP_STATUSES:
            return f"IBPSTATUS: Invalid value '{value}' – must be 'IBP' or blank"
        return None


# ─────────────────────────────────────────────
#  VALIDATION ORCHESTRATOR
# ─────────────────────────────────────────────

class ProductHierarchyValidator:

    def __init__(self):
        self.validators: list[FieldValidator] = self._build_validators()

    def _build_validators(self) -> list[FieldValidator]:
        validators = [
            NotBlankValidator(f, "For FERT/HAWA materials – Field must not be blank")
            for f in NOT_BLANK_FIELDS
        ]
        validators.append(MaterialTypeValidator("MATERIALTYPE", "Field must be FERT or HAWA and must not be blank"))
        validators.append(IBPStatusValidator("IBPSTATUS",       "Field values must be either 'IBP' or blank"))
        return validators

    def validate_row(self, row: pd.Series) -> list[str]:
        errors = []
        for v in self.validators:
            if v.field in row.index or v.field in ("MATERIALTYPE", "IBPSTATUS"):
                result = v.validate(row)
                if result:
                    errors.append(result)
        return errors

    def validate_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()
        df.columns = df.columns.str.strip().str.upper()

        error_list        = []
        error_fields_list = []   # list of sets — which fields errored per row
        error_detail_list = []   # list of dicts — { field: reason } per row

        for _, row in df.iterrows():
            row_errors = self.validate_row(row)

            # Full pipe-separated string for ERROR_COLUMNS in main sheet
            error_list.append(" | ".join(row_errors) if row_errors else "")

            fields_in_error = set()
            field_reason    = {}   # { field_name: reason_string }
            for e in row_errors:
                field = e.split(":")[0].strip()
                fields_in_error.add(field)
                field_reason[field] = e   # full reason string keyed by field

            error_fields_list.append(fields_in_error)
            error_detail_list.append(field_reason)

        # ERROR_COLUMNS replaces VALIDATION_ERRORS
        df["ERROR_COLUMNS"] = error_list
        df["_ERROR_FIELDS"] = error_fields_list    # internal — removed before writing
        df["_ERROR_DETAIL"] = error_detail_list    # internal — removed before writing
        return df


# ─────────────────────────────────────────────
#  EXCEL REPORT BUILDER
# ─────────────────────────────────────────────

class ExcelReportBuilder:

    def __init__(self, df_validated: pd.DataFrame, output_path: str):
        self.df           = df_validated
        self.output       = output_path
        self.wb           = openpyxl.Workbook()
        self.wb.remove(self.wb.active)

        self.error_df     = df_validated[df_validated["ERROR_COLUMNS"] != ""].copy()
        self.error_fields = df_validated["_ERROR_FIELDS"]
        self.error_detail = df_validated["_ERROR_DETAIL"]   # { field: reason } per row

    # ── public entry point ──────────────────────────────

    def build(self):
        self._write_main_sheet()
        # ════════════════════════════════════════
        # All_Errors sheet (COMMENTED OUT)
        # ════════════════════════════════════════
        # self._write_all_errors_sheet()
        # ════════════════════════════════════════
        self._write_summary_sheet()
        self._write_ruleset_sheet()
        self._write_per_field_error_sheets()
        self.wb.save(self.output)
        print(f"\n✅  Report saved → {self.output}")

    # ── Sheet 1: PRODUCTHIERARCHY_FG ────────────────────

    def _write_main_sheet(self):
        ws         = self.wb.create_sheet("PRODUCTHIERARCHY_FG")
        # Drop both internal columns before display
        display_df = self.df.drop(columns=["_ERROR_FIELDS", "_ERROR_DETAIL"])
        headers    = list(display_df.columns)

        ws.append(headers)
        style_header_row(ws, 1, len(headers))
        ws.freeze_panes = "A2"

        for r_idx, (_, row) in enumerate(display_df.iterrows(), start=2):
            has_error      = bool(self.error_fields.iloc[r_idx - 2])
            errored_fields = self.error_fields.iloc[r_idx - 2]

            for c_idx, col in enumerate(headers, start=1):
                cell           = ws.cell(row=r_idx, column=c_idx, value=row[col])
                cell.font      = BODY_FONT
                cell.border    = THIN_BORDER
                cell.alignment = Alignment(vertical="center")

                if col in errored_fields:
                    cell.fill = RED_FILL
                    cell.font = ERR_FONT
                elif has_error:
                    cell.fill = ROW_ERROR_FILL
                else:
                    cell.fill = WHITE_FILL

        auto_width(ws)
        ws.row_dimensions[1].height = 30

    # ════════════════════════════════════════
    # All_Errors sheet (COMMENTED OUT)
    # ════════════════════════════════════════
    # def _write_all_errors_sheet(self):
    #     ws         = self.wb.create_sheet("All_Errors")
    #     display_df = self.error_df.drop(columns=["_ERROR_FIELDS", "_ERROR_DETAIL"])
    #     if display_df.empty:
    #         ws.append(["No errors found"])
    #         return
    #     headers = list(display_df.columns)
    #     ws.append(headers)
    #     style_header_row(ws, 1, len(headers))
    #     ws.freeze_panes = "A2"
    #     for r_idx, (_, row) in enumerate(display_df.iterrows(), start=2):
    #         errored_fields = self.error_fields.loc[row.name]
    #         for c_idx, col in enumerate(headers, start=1):
    #             cell = ws.cell(row=r_idx, column=c_idx, value=row[col])
    #             cell.border = THIN_BORDER
    #             cell.font   = BODY_FONT
    #             cell.alignment = Alignment(vertical="center")
    #             if col in errored_fields:
    #                 cell.fill = RED_FILL
    #                 cell.font = ERR_FONT
    #             else:
    #                 cell.fill = ROW_ERROR_FILL
    #     auto_width(ws)
    # ════════════════════════════════════════

    # ── Summary sheet ───────────────────────────────────

    def _write_summary_sheet(self):
        ws = self.wb.create_sheet("Summary")

        total_rows          = len(self.df)
        records_with_errors = len(self.error_df)
        records_passing     = total_rows - records_with_errors

        field_counts: dict[str, int] = defaultdict(int)
        for fields_set in self.error_fields:
            for f in fields_set:
                field_counts[f] += 1

        # ── Title ──
        title_cell           = ws.cell(row=1, column=1, value="ProductHierarchy FG Validation Summary")
        title_cell.font      = Font(name="Arial", bold=True, size=14)
        title_cell.fill      = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells("A1:D1")
        ws.row_dimensions[1].height = 24

        # ── Column headers — % of Total Records COMMENTED OUT ──
        headers = [
            "#",
            "Field Name",
            "Error Count",
            # "% of Total Records",   # <-- COMMENTED OUT
        ]
        for c_idx, h in enumerate(headers, start=1):
            cell           = ws.cell(row=3, column=c_idx, value=h)
            cell.fill      = TITLE_FILL
            cell.font      = Font(name="Arial", bold=True)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # ── Per-field rows ──
        row_num = 4
        for field_num, (col_name, count) in enumerate(sorted(field_counts.items()), start=1):
            # pct = f"{(count / total_rows * 100):.2f}%"   # COMMENTED OUT

            ws.cell(row=row_num, column=1, value=field_num).font = BODY_FONT
            ws.cell(row=row_num, column=2, value=col_name).font  = BODY_FONT
            ws.cell(row=row_num, column=3, value=count).font     = BODY_FONT
            # ws.cell(row=row_num, column=4, value=pct).font     = BODY_FONT   # COMMENTED OUT

            for c in range(1, 4):   # was range(1, 5)
                ws.cell(row=row_num, column=c).border    = THIN_BORDER
                ws.cell(row=row_num, column=c).alignment = Alignment(horizontal="center")

            row_num += 1

        # ── TOTAL row ──
        total_errors = sum(field_counts.values())
        # total_pct  = f"{(total_errors / total_rows * 100):.2f}%"   # COMMENTED OUT

        ws.cell(row=row_num, column=2, value="TOTAL").font      = Font(name="Arial", bold=True)
        ws.cell(row=row_num, column=3, value=total_errors).font = Font(name="Arial", bold=True)
        # ws.cell(row=row_num, column=4, value=total_pct).font  = Font(name="Arial", bold=True)   # COMMENTED OUT
        for c in range(1, 4):   # was range(1, 5)
            ws.cell(row=row_num, column=c).fill      = TOTAL_FILL
            ws.cell(row=row_num, column=c).border    = THIN_BORDER
            ws.cell(row=row_num, column=c).alignment = Alignment(horizontal="center")

        # ── Spacer then stats block ──
        row_num += 2

        STATS_LABEL_FILL = PatternFill("solid", start_color="EDEDED", end_color="EDEDED")
        for label, value in [
            ("Total Records:",       total_rows),
            ("Records with Errors:", records_with_errors),
            ("Records Passing:",     records_passing),
        ]:
            label_cell           = ws.cell(row=row_num, column=1, value=label)
            label_cell.font      = Font(name="Arial", bold=True, size=10)
            label_cell.fill      = STATS_LABEL_FILL
            label_cell.border    = THIN_BORDER
            label_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)

            value_cell           = ws.cell(row=row_num, column=3, value=value)
            value_cell.font      = BODY_FONT
            value_cell.border    = THIN_BORDER
            value_cell.alignment = Alignment(horizontal="center", vertical="center")
            row_num += 1

        for c_idx, width in enumerate([6, 34, 16], start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = width

    # ── Rule_Set sheet (same template as Part & Site) ────

    def _write_ruleset_sheet(self):
        ws = self.wb.create_sheet("Rule_Set")

        # ── Title row (same as Part/Site) ──
        title_cell           = ws.cell(row=1, column=1, value="ProductHierarchy FG – Validation Rules")
        title_cell.font      = Font(name="Arial", bold=True, size=13)
        title_cell.fill      = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="center")
        ws.merge_cells("A1:C1")
        ws.row_dimensions[1].height = 22

        # ── Column headers ──
        for c_idx, h in enumerate(["#", "Field", "Rule Description"], start=1):
            cell           = ws.cell(row=3, column=c_idx, value=h)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        current_row = 4
        rule_num    = 1

        for field, rules_list in RULES_CONTENT.items():
            num_rules = len(rules_list)

            for r_idx, rule_text in enumerate(rules_list):
                # Rule number cell
                num_cell           = ws.cell(row=current_row, column=1, value=rule_num if r_idx == 0 else "")
                num_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                num_cell.fill      = RULE_FILL
                num_cell.border    = THIN_BORDER
                num_cell.alignment = Alignment(horizontal="center", vertical="center")

                # Field name cell
                field_cell           = ws.cell(row=current_row, column=2, value=field if r_idx == 0 else "")
                field_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                field_cell.fill      = RULE_FILL
                field_cell.border    = THIN_BORDER
                field_cell.alignment = Alignment(vertical="center")

                # Rule description cell
                desc_cell           = ws.cell(row=current_row, column=3, value=rule_text)
                desc_cell.font      = BODY_FONT
                desc_cell.border    = THIN_BORDER
                desc_cell.alignment = Alignment(wrap_text=True, vertical="center")

                current_row += 1

            # Merge # and Field cells vertically for multi-rule fields
            if num_rules > 1:
                s = current_row - num_rules
                e = current_row - 1
                ws.merge_cells(start_row=s, start_column=1, end_row=e, end_column=1)
                ws.merge_cells(start_row=s, start_column=2, end_row=e, end_column=2)

            rule_num += 1

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 65

    # ── Per-field error sheets ───────────────────────────

    def _write_per_field_error_sheets(self):
        if self.error_df.empty:
            return

        field_rows: dict[str, list] = defaultdict(list)
        for idx, row in self.error_df.iterrows():
            for field in self.error_fields.loc[idx]:
                field_rows[field].append(idx)

        # Display columns — drop both internal cols
        display_cols = [c for c in self.df.columns if c not in ("_ERROR_FIELDS", "_ERROR_DETAIL")]

        for field, row_indices in sorted(field_rows.items()):
            sheet_name = field[:28] + "_ERR" if len(field) > 28 else field + "_ERR"
            existing   = [s.title for s in self.wb.worksheets]
            counter    = 1
            base_name  = sheet_name
            while sheet_name in existing:
                sheet_name = f"{base_name[:25]}_{counter}"
                counter   += 1

            ws = self.wb.create_sheet(sheet_name)

            # Build subset with only rows that failed THIS specific field
            subset = self.df.loc[row_indices, display_cols].copy()

            # Override ERROR_COLUMNS — show ONLY this field's specific reason
            subset["ERROR_COLUMNS"] = subset.index.map(
                lambda i, f=field: self.error_detail.loc[i].get(f, "")
            )

            final_cols = list(subset.columns)
            ws.append(final_cols)
            style_header_row(ws, 1, len(final_cols))
            ws.freeze_panes = "A2"

            col_idx_map = {col: i for i, col in enumerate(final_cols, start=1)}

            for r_idx, (orig_idx, row) in enumerate(subset.iterrows(), start=2):
                # Step 1: whole row yellow
                for c_idx, col in enumerate(final_cols, start=1):
                    cell           = ws.cell(row=r_idx, column=c_idx, value=row[col])
                    cell.font      = BODY_FONT
                    cell.border    = THIN_BORDER
                    cell.alignment = Alignment(vertical="center")
                    cell.fill      = ROW_ERROR_FILL   # whole row yellow

                # Step 2: only THIS field's cell goes red
                if field in col_idx_map:
                    target_cell      = ws.cell(row=r_idx, column=col_idx_map[field])
                    target_cell.fill = RED_FILL
                    target_cell.font = ERR_FONT

            # Row count note at bottom
            note_row = len(row_indices) + 3
            ws.cell(
                row=note_row, column=1,
                value=f"Total error rows for '{field}': {len(row_indices)}",
            ).font = Font(name="Arial", italic=True, size=9, bold=True)

            auto_width(ws)


# ─────────────────────────────────────────────
#  PIPELINE CONTROLLER
# ─────────────────────────────────────────────

class ValidationPipeline:

    def __init__(self, input_path: str, output_path: str):
        self.input_path  = input_path
        self.output_path = output_path
        self.validator   = ProductHierarchyValidator()

    def run(self):
        print(f"📂  Reading: {self.input_path}")
        df = self._read_input()

        print(f"🔍  Validating {len(df)} rows …")
        df_validated = self.validator.validate_dataframe(df)

        error_count = (df_validated["ERROR_COLUMNS"] != "").sum()
        print(f"⚠️   Errors found in {error_count} row(s)")

        print("📝  Building Excel report …")
        builder = ExcelReportBuilder(df_validated, self.output_path)
        builder.build()

    def _read_input(self) -> pd.DataFrame:
        ext = os.path.splitext(self.input_path)[1].lower()
        if ext in (".xlsx", ".xlsm", ".xls"):
            return pd.read_excel(self.input_path, dtype=str)
        elif ext == ".csv":
            return pd.read_csv(self.input_path, dtype=str)
        else:
            raise ValueError(f"Unsupported file type: {ext}")


# ─────────────────────────────────────────────
#  ENTRY POINT
# ─────────────────────────────────────────────

if __name__ == "__main__":
    pipeline = ValidationPipeline(INPUT_FILE, OUTPUT_FILE)
    pipeline.run()

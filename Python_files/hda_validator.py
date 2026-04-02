"""
Historical Demand Actuals (HDA) - Excel Data Validation Tool
Validates HDA table fields against Part, Site, and Customer master data.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import re


# ─────────────────────────────────────────────
#  FILE PATHS  –  update these
# ─────────────────────────────────────────────
HDA_INPUT_FILE      = r"D:\SEM-8\Data Rules Set Check\Data_rulesets_check\Excel_Files\HDA.xlsx"
PART_INPUT_FILE     = r"D:\SEM-8\Data Rules Set Check\Data_rulesets_check\Excel_Files\Part.xlsx"
SITE_INPUT_FILE     = r"D:\SEM-8\Data Rules Set Check\Data_rulesets_check\Excel_Files\Site.xlsx"
CUSTOMER_INPUT_FILE = r"D:\SEM-8\Data Rules Set Check\Data_rulesets_check\Excel_Files\Customer.xlsx"
OUTPUT_FILE         = r"D:\SEM-8\Data Rules Set Check\Data_rulesets_check\Validated_HDA.xlsx"


# ─────────────────────────────────────────────
#  Colours  (aligned with Part / Site / PH-FG)
# ─────────────────────────────────────────────
RED_FILL    = PatternFill("solid", start_color="FF0000", end_color="FF0000")
ROW_FILL    = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")   # light yellow
HDR_FILL    = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")   # light blue
RULE_FILL   = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")   # light green
TITLE_FILL  = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")   # header blue
WHITE_FILL  = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
TOTAL_FILL  = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
STATS_FILL  = PatternFill("solid", start_color="EDEDED", end_color="EDEDED")

HDR_FONT    = Font(bold=True, name="Arial")
BODY_FONT   = Font(name="Arial", size=10)
ERR_FONT    = Font(name="Arial", size=10, bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

YYYYMMDD_RE = re.compile(r"^\d{8}$")


# ─────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────

def is_blank(value) -> bool:
    if value is None:
        return True
    try:
        import math
        if isinstance(value, float) and math.isnan(value):
            return True
    except Exception:
        pass
    return str(value).strip() == ""


def clean(value) -> str:
    return "" if is_blank(value) else str(value).strip()


def style_header(ws, row: int, num_cols: int):
    for c in range(1, num_cols + 1):
        cell           = ws.cell(row=row, column=c)
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER


def auto_width(ws, min_w=10, max_w=55):
    for col in ws.columns:
        length = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)


def is_valid_yyyymmdd(value) -> bool:
    """Check YYYYMMDD format and calendar validity."""
    s = clean(value)
    if not YYYYMMDD_RE.match(s):
        return False
    try:
        pd.to_datetime(s, format="%Y%m%d")
        return True
    except Exception:
        return False


# ═══════════════════════════════════════════════
#  MASTER DATA LOADER
# ═══════════════════════════════════════════════

class MasterDataLoader:
    """Loads Part, Site, and Customer reference data into lookup sets."""

    def __init__(self, part_path: str, site_path: str, customer_path: str):
        self.part_path     = part_path
        self.site_path     = site_path
        self.customer_path = customer_path

        # Populated by load()
        self.part_combinations: set   = set()   # (MATERIALNUMBER, PLANT)
        self.site_plants: set         = set()   # PLANT values from Site
        self.customer_combinations: set = set() # (SUPPLYINGPLANT, CUSTOMER)

    def load(self):
        # ── Part master ──
        part_df = pd.read_excel(self.part_path, dtype=str)
        part_df.columns = part_df.columns.str.strip().str.upper()
        if "MATERIALNUMBER" not in part_df.columns or "PLANT" not in part_df.columns:
            raise ValueError("Part file must have MATERIALNUMBER and PLANT columns.")
        for _, row in part_df.iterrows():
            mat   = clean(row.get("MATERIALNUMBER"))
            plant = clean(row.get("PLANT"))
            if mat and plant:
                self.part_combinations.add((mat, plant))
        print(f"    Part combinations loaded     : {len(self.part_combinations)}")

        # ── Site master ──
        site_df = pd.read_excel(self.site_path, dtype=str)
        site_df.columns = site_df.columns.str.strip().str.upper()
        if "PLANT" not in site_df.columns:
            raise ValueError("Site file must have a PLANT column.")
        self.site_plants = set(site_df["PLANT"].dropna().str.strip().tolist())
        print(f"    Site plants loaded           : {len(self.site_plants)}")

        # ── Customer master ──
        cust_df = pd.read_excel(self.customer_path, dtype=str)
        cust_df.columns = cust_df.columns.str.strip().str.upper()
        if "SUPPLYINGPLANT" not in cust_df.columns or "CUSTOMER" not in cust_df.columns:
            raise ValueError("Customer file must have SUPPLYINGPLANT and CUSTOMER columns.")
        for _, row in cust_df.iterrows():
            sp   = clean(row.get("SUPPLYINGPLANT"))
            cust = clean(row.get("CUSTOMER"))
            if sp and cust:
                self.customer_combinations.add((sp, cust))
        print(f"    Customer combinations loaded : {len(self.customer_combinations)}")


# ═══════════════════════════════════════════════
#  RULE ENGINE  (returns reason string or "")
# ═══════════════════════════════════════════════

class HDAValidator:
    """
    Each validate_* method returns:
      ""            → PASS
      reason string → FAIL  (written verbatim into ERROR_COLUMNS)
    """

    # Rules sheet content
    RULES_CONTENT = {
        "MATERIAL": [
            "Must not be blank.",
            "The combination of MATERIAL + PLANT in HDA must exist as MATERIALNUMBER + PLANT in the Part master table.",
        ],
        "PLANT": [
            "Must not be blank.",
            "PLANT value must exist in the PLANT column of the Part master table (checked via MATERIAL+PLANT combo).",
            "PLANT value must also exist in the PLANT column of the Site master table.",
        ],
        "SOLDTOPARTY": [
            "Must not be blank.",
            "The combination of PLANT (HDA) + SOLDTOPARTY (HDA) must exist as SUPPLYINGPLANT + CUSTOMER in the Customer master table.",
        ],
        "BILLING_WEEK_START": [
            "Must not be blank.",
            "Must be in YYYYMMDD format (8 digits, valid calendar date).",
        ],
    }

    def __init__(self, master: MasterDataLoader):
        self.master = master

    # ── Rule 1: MATERIAL ──────────────────────────────
    def validate_material(self, row) -> str:
        mat   = clean(row.get("MATERIAL"))
        plant = clean(row.get("PLANT"))

        if not mat:
            return "MATERIAL: Field is blank — material number is mandatory"

        # PLANT blank check is handled in validate_plant; here just check the combo
        if plant and (mat, plant) not in self.master.part_combinations:
            return (
                f"MATERIAL: Combination MATERIAL='{mat}' + PLANT='{plant}' "
                f"does not exist in Part master (MATERIALNUMBER + PLANT)"
            )
        return ""

    # ── Rule 2: PLANT ─────────────────────────────────
    def validate_plant(self, row) -> str:
        plant = clean(row.get("PLANT"))
        mat   = clean(row.get("MATERIAL"))

        if not plant:
            return "PLANT: Field is blank — plant code is mandatory"

        # Check Part master (via combo — if mat present, combo already checked in MATERIAL rule;
        # here we verify the plant itself appears in at least one Part combo)
        part_plants = {p for (_, p) in self.master.part_combinations}
        if plant not in part_plants:
            return (
                f"PLANT: '{plant}' does not appear in the PLANT column "
                f"of the Part master table"
            )

        # Check Site master
        if plant not in self.master.site_plants:
            return (
                f"PLANT: '{plant}' does not exist in the PLANT column "
                f"of the Site master table"
            )

        return ""

    # ── Rule 3: SOLDTOPARTY ───────────────────────────
    def validate_soldtoparty(self, row) -> str:
        soldto = clean(row.get("SOLDTOPARTY"))
        plant  = clean(row.get("PLANT"))

        if not soldto:
            return "SOLDTOPARTY: Field is blank — sold-to party is mandatory"

        if plant and (plant, soldto) not in self.master.customer_combinations:
            return (
                f"SOLDTOPARTY: Combination PLANT='{plant}' + SOLDTOPARTY='{soldto}' "
                f"does not exist in Customer master (SUPPLYINGPLANT + CUSTOMER)"
            )
        return ""

    # ── Rule 4: BILLING_WEEK_START ────────────────────
    def validate_billing_week_start(self, row) -> str:
        val = clean(row.get("BILLING_WEEK_START"))

        if not val:
            return "BILLING_WEEK_START: Field is blank — billing week start date is mandatory"

        if not is_valid_yyyymmdd(val):
            return (
                f"BILLING_WEEK_START: '{val}' is not a valid YYYYMMDD date "
                f"— must be 8 digits representing a real calendar date"
            )
        return ""

    def get_rules(self) -> dict:
        return {
            "MATERIAL":           self.validate_material,
            "PLANT":              self.validate_plant,
            "SOLDTOPARTY":        self.validate_soldtoparty,
            "BILLING_WEEK_START": self.validate_billing_week_start,
        }


# ═══════════════════════════════════════════════
#  VALIDATOR  (builds error_map + reason_map)
# ═══════════════════════════════════════════════

class HDATableValidator:

    def __init__(self, hda_path: str, master: MasterDataLoader):
        self.hda_path   = hda_path
        self.master     = master
        self.df         = pd.DataFrame()
        self.error_map  = {}   # { row_idx: [col, ...] }
        self.reason_map = {}   # { row_idx: [reason_str, ...] }

    def load(self):
        self.df = pd.read_excel(self.hda_path, dtype=str)
        self.df.columns = self.df.columns.str.strip().str.upper()
        print(f"    HDA columns detected : {list(self.df.columns)}")

    def validate(self):
        engine = HDAValidator(self.master)
        rules  = engine.get_rules()

        for idx, row in self.df.iterrows():
            failed_cols    = []
            failed_reasons = []

            for col, rule_fn in rules.items():
                if col not in self.df.columns:
                    continue
                try:
                    reason = rule_fn(row)
                except Exception as e:
                    reason = f"{col}: Unexpected validation error — {e}"

                if reason:
                    failed_cols.append(col)
                    failed_reasons.append(reason)

            if failed_cols:
                self.error_map[idx]  = failed_cols
                self.reason_map[idx] = failed_reasons

    def get_error_series(self) -> pd.Series:
        """Pipe-separated, condition-specific reason strings per row."""
        return pd.Series(
            {idx: " | ".join(reasons) for idx, reasons in self.reason_map.items()},
            dtype=str,
        )

    def get_errors_by_field(self) -> dict:
        field_errors: dict = {}
        for row_idx, bad_cols in self.error_map.items():
            for col in bad_cols:
                field_errors.setdefault(col, []).append(row_idx)
        return field_errors


# ═══════════════════════════════════════════════
#  REPORT WRITER
# ═══════════════════════════════════════════════

class HDAReportWriter:

    SHEET_MAIN    = "HDA"
    SHEET_SUMMARY = "Summary"
    SHEET_RULES   = "Rules"

    def __init__(self, validator: HDATableValidator, output_path: str):
        self.validator   = validator
        self.output_path = output_path

    # ── shared helpers ────────────────────────

    def _write_header(self, ws, columns):
        for c_idx, col_name in enumerate(columns, start=1):
            cell           = ws.cell(row=1, column=c_idx, value=col_name)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = THIN_BORDER

    def _write_rows(self, ws, df: pd.DataFrame):
        for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
            for c_idx, value in enumerate(row, start=1):
                cell           = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.font      = BODY_FONT
                cell.alignment = Alignment(vertical="center")

    def _highlight(self, ws, df: pd.DataFrame, error_map: dict, col_index: dict):
        for df_idx, bad_cols in error_map.items():
            excel_row = df_idx + 2
            for c in range(1, len(df.columns) + 1):
                ws.cell(row=excel_row, column=c).fill = ROW_FILL
            for col_name in bad_cols:
                if col_name in col_index:
                    cell      = ws.cell(row=excel_row, column=col_index[col_name])
                    cell.fill = RED_FILL
                    cell.font = ERR_FONT

    def _set_widths(self, ws):
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    # ── Sheet 1: HDA Full Data ────────────────

    def _write_main_sheet(self, wb, df: pd.DataFrame):
        ws = wb.active
        ws.title = self.SHEET_MAIN

        self._write_header(ws, df.columns)
        self._write_rows(ws, df)

        col_index = {col: i for i, col in enumerate(df.columns, start=1)}
        self._highlight(ws, df, self.validator.error_map, col_index)

        # Rows with no error → plain white
        error_rows = set(self.validator.error_map.keys())
        for df_idx in range(len(df)):
            if df_idx not in error_rows:
                for c in range(1, len(df.columns) + 1):
                    ws.cell(row=df_idx + 2, column=c).fill = WHITE_FILL

        ws.freeze_panes = "A2"
        self._set_widths(ws)
        ws.row_dimensions[1].height = 28

    # ════════════════════════════════════════
    # Error Rows sheet (COMMENTED OUT)
    # ════════════════════════════════════════
    # def _write_error_rows_sheet(self, wb, df: pd.DataFrame):
    #     error_df    = df[df.index.isin(self.validator.error_map.keys())].copy()
    #     ws          = wb.create_sheet("Error Rows")
    #     col_index   = {col: i for i, col in enumerate(error_df.columns, start=1)}
    #     self._write_header(ws, error_df.columns)
    #     self._write_rows(ws, error_df)
    #     for sheet_row, orig_idx in enumerate(error_df.index, start=2):
    #         bad_cols = self.validator.error_map[orig_idx]
    #         for c in range(1, len(error_df.columns) + 1):
    #             ws.cell(row=sheet_row, column=c).fill = ROW_FILL
    #         for col_name in bad_cols:
    #             if col_name in col_index:
    #                 cell      = ws.cell(row=sheet_row, column=col_index[col_name])
    #                 cell.fill = RED_FILL
    #                 cell.font = ERR_FONT
    #     self._set_widths(ws)
    # ════════════════════════════════════════

    # ── Sheet 2: Summary ──────────────────────

    def _write_summary_sheet(self, wb, total_rows: int):
        ws          = wb.create_sheet(self.SHEET_SUMMARY)
        error_map   = self.validator.error_map

        # ── Title ──
        title_cell           = ws.cell(row=1, column=1, value="Historical Demand Actuals Validation Summary")
        title_cell.font      = Font(name="Arial", bold=True, size=14)
        title_cell.fill      = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells("A1:E1")
        ws.row_dimensions[1].height = 24

        # ── Column headers ──
        for c_idx, h in enumerate(["#", "Field Name", "Error Count", "% of Total Records"], start=1):
            cell           = ws.cell(row=3, column=c_idx, value=h)
            cell.fill      = TITLE_FILL
            cell.font      = Font(name="Arial", bold=True)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # ── Per-field rows ──
        col_error_counts: dict = {}
        for bad_cols in error_map.values():
            for col in bad_cols:
                col_error_counts[col] = col_error_counts.get(col, 0) + 1

        row_num = 4
        for field_num, (col_name, count) in enumerate(col_error_counts.items(), start=1):
            pct = f"{(count / total_rows * 100):.2f}%" if total_rows > 0 else "0.00%"

            ws.cell(row=row_num, column=1, value=field_num).font = BODY_FONT
            ws.cell(row=row_num, column=2, value=col_name).font  = BODY_FONT
            ws.cell(row=row_num, column=3, value=count).font     = BODY_FONT
            ws.cell(row=row_num, column=4, value=pct).font       = BODY_FONT

            for c in range(1, 5):
                ws.cell(row=row_num, column=c).border    = THIN_BORDER
                ws.cell(row=row_num, column=c).alignment = Alignment(horizontal="center")

            row_num += 1

        # ── TOTAL row ──
        total_errors = sum(col_error_counts.values())
        total_pct    = f"{(total_errors / total_rows * 100):.2f}%" if total_rows > 0 else "0.00%"

        ws.cell(row=row_num, column=2, value="TOTAL").font      = Font(name="Arial", bold=True)
        ws.cell(row=row_num, column=3, value=total_errors).font = Font(name="Arial", bold=True)
        ws.cell(row=row_num, column=4, value=total_pct).font    = Font(name="Arial", bold=True)
        for c in range(1, 5):
            ws.cell(row=row_num, column=c).fill      = TOTAL_FILL
            ws.cell(row=row_num, column=c).border    = THIN_BORDER
            ws.cell(row=row_num, column=c).alignment = Alignment(horizontal="center")

        # ── Stats block ──
        row_num += 2
        records_with_errors = len(error_map)
        records_passing     = total_rows - records_with_errors

        for label, value in [
            ("Total Records:",       total_rows),
            ("Records with Errors:", records_with_errors),
            ("Records Passing:",     records_passing),
        ]:
            label_cell           = ws.cell(row=row_num, column=1, value=label)
            label_cell.font      = Font(name="Arial", bold=True, size=10)
            label_cell.fill      = STATS_FILL
            label_cell.border    = THIN_BORDER
            label_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)

            value_cell           = ws.cell(row=row_num, column=3, value=value)
            value_cell.font      = BODY_FONT
            value_cell.border    = THIN_BORDER
            value_cell.alignment = Alignment(horizontal="center", vertical="center")
            row_num += 1

        # ── Column widths ──
        for c_idx, width in enumerate([6, 30, 16, 22], start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = width

    # ── Per-field error sheets ────────────────

    def _write_field_error_sheets(self, wb, df: pd.DataFrame):
        field_errors = self.validator.get_errors_by_field()

        for field_name, row_indices in field_errors.items():
            sheet_name = field_name[:31].replace("/", "-").replace("\\", "-").replace("*", "")
            ws         = wb.create_sheet(sheet_name)
            subset     = df.loc[row_indices].copy()

            self._write_header(ws, subset.columns)
            col_idx_map = {col: i for i, col in enumerate(subset.columns, start=1)}

            for excel_row, (_, row_data) in enumerate(subset.iterrows(), start=2):
                for c_idx, value in enumerate(row_data, start=1):
                    cell           = ws.cell(row=excel_row, column=c_idx, value=value)
                    cell.font      = BODY_FONT
                    cell.fill      = ROW_FILL
                    cell.alignment = Alignment(vertical="center")

                if field_name in col_idx_map:
                    bad_cell      = ws.cell(row=excel_row, column=col_idx_map[field_name])
                    bad_cell.fill = RED_FILL
                    bad_cell.font = ERR_FONT

            self._set_widths(ws)

            note_row = len(subset) + 3
            ws.cell(
                row=note_row, column=1,
                value=f"Total error rows for '{field_name}': {len(subset)}",
            ).font = Font(name="Arial", italic=True, size=9, bold=True)

    # ── Rules sheet ───────────────────────────

    def _write_rules_sheet(self, wb):
        ws = wb.create_sheet(self.SHEET_RULES)

        title_cell           = ws.cell(row=1, column=1, value="Historical Demand Actuals – Validation Rules")
        title_cell.font      = Font(name="Arial", bold=True, size=13)
        title_cell.fill      = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="center")
        ws.merge_cells("A1:C1")
        ws.row_dimensions[1].height = 22

        for c_idx, h in enumerate(["#", "Field", "Rule Description"], start=1):
            cell           = ws.cell(row=3, column=c_idx, value=h)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        current_row = 4
        rule_num    = 1

        for field, rules_list in HDAValidator.RULES_CONTENT.items():
            num_rules = len(rules_list)

            for r_idx, rule_text in enumerate(rules_list):
                num_cell           = ws.cell(row=current_row, column=1, value=rule_num if r_idx == 0 else "")
                num_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                num_cell.fill      = RULE_FILL
                num_cell.border    = THIN_BORDER
                num_cell.alignment = Alignment(horizontal="center", vertical="center")

                field_cell           = ws.cell(row=current_row, column=2, value=field if r_idx == 0 else "")
                field_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                field_cell.fill      = RULE_FILL
                field_cell.border    = THIN_BORDER
                field_cell.alignment = Alignment(vertical="center")

                desc_cell           = ws.cell(row=current_row, column=3, value=rule_text)
                desc_cell.font      = BODY_FONT
                desc_cell.border    = THIN_BORDER
                desc_cell.alignment = Alignment(wrap_text=True, vertical="center")

                current_row += 1

            if num_rules > 1:
                s = current_row - num_rules
                e = current_row - 1
                ws.merge_cells(start_row=s, start_column=1, end_row=e, end_column=1)
                ws.merge_cells(start_row=s, start_column=2, end_row=e, end_column=2)

            rule_num += 1

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 24
        ws.column_dimensions["C"].width = 72

    # ── Main write ────────────────────────────

    def write(self):
        v   = self.validator
        df  = v.df.copy()

        # ERROR_COLUMNS: condition-specific, pipe-separated reason strings
        error_series        = v.get_error_series()
        df["ERROR_COLUMNS"] = df.index.map(lambda i: error_series.get(i, ""))

        wb = openpyxl.Workbook()

        self._write_main_sheet(wb, df)
        # Error Rows sheet commented out — see method above
        self._write_summary_sheet(wb, total_rows=len(df))
        self._write_field_error_sheets(wb, df)
        self._write_rules_sheet(wb)

        wb.save(self.output_path)
        print(f"\n✅  Output saved  → {self.output_path}")
        print(f"   Total rows    : {len(df)}")
        print(f"   Error rows    : {len(v.error_map)}")
        print(f"   Field sheets  : {list(v.get_errors_by_field().keys())}")


# ═══════════════════════════════════════════════
#  ORCHESTRATOR
# ═══════════════════════════════════════════════

class HDAProcessor:

    def __init__(self, hda_path, part_path, site_path, customer_path, output_path):
        self.master    = MasterDataLoader(part_path, site_path, customer_path)
        self.validator = HDATableValidator(hda_path, self.master)
        self.writer    = HDAReportWriter(self.validator, output_path)

    def run(self):
        print("📂  Loading master data …")
        self.master.load()
        print("📂  Loading HDA file …")
        self.validator.load()
        print("🔍  Validating rules …")
        self.validator.validate()
        print("📝  Writing report …")
        self.writer.write()


# ═══════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════

if __name__ == "__main__":
    processor = HDAProcessor(
        hda_path      = HDA_INPUT_FILE,
        part_path     = PART_INPUT_FILE,
        site_path     = SITE_INPUT_FILE,
        customer_path = CUSTOMER_INPUT_FILE,
        output_path   = OUTPUT_FILE,
    )
    processor.run()
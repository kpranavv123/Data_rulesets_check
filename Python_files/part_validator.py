import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
#  FILE PATHS  –  update these
# ─────────────────────────────────────────────
INPUT_FILE  = r"D:\SEM-8\Data Rules Set Check\Data_rulesets_check\Excel_Files\Part.xlsx"
OUTPUT_FILE = r"D:\SEM-8\Data Rules Set Check\Data_rulesets_check\Validated_Part.xlsx"


# ─────────────────────────────────────────────
#  CONSOLIDATED PL LIST
# ─────────────────────────────────────────────
VALID_PLANTS = [
    "1127", "1100", "1105", "1146", "1156", "1107", "1157", "1158", "1166", "1180",
    "1184", "1186", "1197", "1203", "1204", "1211", "1213", "1214", "1218", "1223",
    "1110", "1225", "1226", "1229", "1233", "1234", "1240", "1113", "1248", "1253",
    "1257", "1258", "1265", "1114", "1145", "1275", "1279", "1416", "1421", "1423",
    "1425", "1426", "1428", "1429", "1430", "1432", "1433", "1436", "1437", "1438",
    "1439", "1440", "1442", "1445", "1449", "1451", "1452", "1455", "1463", "1471",
    "1473", "1475", "1476", "1477", "1478", "1480", "1481", "1483", "1484", "1485",
    "1487", "1488", "1491", "1495", "1500", "1501", "1505", "1506", "1507", "1508",
    "1509", "1521", "1525", "1563", "1578", "1650", "1651", "1652", "1654", "1656",
    "1657", "1658", "1659", "1661", "1579", "1627", "1589", "1623", "1646", "1647",
    "1640", "1112", "5011637", "5123296", "5123742", "4007430", "1722", "1724", "1725",
    "1726", "1731", "1732", "1733", "1734", "1738", "1739", "1740", "1742", "1754",
    "1757", "1758", "1771", "1774", "1780", "1784", "1785", "1788", "1511", "1512",
    "1948", "1448", "4011702", "5013796", "5018849", "5011407", "5015073", "5011308",
    "5123531", "1642", "1638", "1104", "1104A", "1643", "1106", "1109", "1645", "1111",
    "1648", "1649", "1653", "1801", "1802", "2082", "2091", "2088", "2089", "2083",
    "2084", "2085", "2090", "2081", "2086", "2087", "1554", "1520", "1472", "1571",
    "1298", "1295", "1196", "1292", "1176", "1441", "1522", "1494", "1489", "1518",
    "1249", "1296", "1137", "1208", "1155", "1569", "1235", "1281", "1503", "1482",
    "1135", "1205", "1241", "1499", "1462", "1555", "1559", "1575", "2092", "1558",
    "1601", "1125", "1256", "1568",
]


# ─────────────────────────────────────────────
#  Colours
# ─────────────────────────────────────────────
RED_FILL   = PatternFill("solid", start_color="FF0000", end_color="FF0000")
ROW_FILL   = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
HDR_FILL   = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
RULE_FILL  = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
TITLE_FILL = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
HDR_FONT   = Font(bold=True, name="Arial")
BODY_FONT  = Font(name="Arial", size=10)
ERR_FONT   = Font(name="Arial", size=10, bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


# ─────────────────────────────────────────────
#  Human-readable one-liner error messages per field
# ─────────────────────────────────────────────
ERROR_MESSAGES = {
    "MATERIALNUMBER":     "MATERIALNUMBER: Must be 14xxxxxxxxxxxxxx (FERT) or 15xxxxxxxxxxxxxx (HAWA) — invalid range or blank",
    "PLANT":              "PLANT: Value is blank or not in the Consolidated PL list",
    "PRODUCTDESCRIPTION": "PRODUCTDESCRIPTION: Field is blank — description is mandatory",
    "PRODUCTTYPE":        "PRODUCTTYPE: Must be FERT or HAWA — invalid or blank value found",
    "PRODUCTHIERARCHY":   "PRODUCTHIERARCHY: Field is blank — hierarchy is mandatory",
    "MRPTYPE":            "MRPTYPE: Must be ND or PD — invalid or blank value found",
    "PROCUREMENTTYPE":    "PROCUREMENTTYPE: Field is blank — procurement type is mandatory",
    "IBPSTATUS":          "IBPSTATUS: Must be 'IBP' or blank — unexpected value found",
    "XPLANTMATSTATUS":    "XPLANTMATSTATUS: Must be '2' or blank — unexpected value found",
}


# ══════════════════════════════════════════════
#  Rule Engine
# ══════════════════════════════════════════════
class RuleEngine:
    """Contains all column-level validation rules for the Part Table."""

    def __init__(self, valid_plants: list):
        self.valid_plants = [str(p).strip() for p in valid_plants]

    @staticmethod
    def _is_blank(value) -> bool:
        return pd.isna(value) or str(value).strip() == ""

    def _check_not_blank(self, value) -> bool:
        return not self._is_blank(value)

    # ─ Rule 1: MATERIALNUMBER ─
    def validate_material_number(self, row) -> bool:
        mat   = row.get("MATERIALNUMBER")
        ptype = str(row.get("PRODUCTTYPE", "")).strip().upper()
        if self._is_blank(mat):
            return False
        try:
            num = int(float(str(mat).strip()))
        except (ValueError, TypeError):
            return False
        if ptype == "FERT":
            return 14000000000000 <= num <= 14999999999999
        if ptype == "HAWA":
            return 15000000000000 <= num <= 15999999999999
        return False

    # ─ Rule 2: PLANT ─
    def validate_plant(self, row) -> bool:
        val = row.get("PLANT")
        if self._is_blank(val):
            return False
        return str(val).strip() in self.valid_plants

    # ─ Rule 3: PRODUCTDESCRIPTION ─
    def validate_product_description(self, row) -> bool:
        return self._check_not_blank(row.get("PRODUCTDESCRIPTION"))

    # ─ Rule 4: PRODUCTTYPE ─
    def validate_product_type(self, row) -> bool:
        val = str(row.get("PRODUCTTYPE", "")).strip().upper()
        return val in {"FERT", "HAWA"}

    # ─ Rule 5: PRODUCTHIERARCHY ─
    def validate_product_hierarchy(self, row) -> bool:
        return self._check_not_blank(row.get("PRODUCTHIERARCHY"))

    # ─ Rule 6: MRPTYPE ─
    def validate_mrp_type(self, row) -> bool:
        val = str(row.get("MRPTYPE", "")).strip().upper()
        return val in {"ND", "PD"}

    # ─ Rule 7: PROCUREMENTTYPE ─
    def validate_procurement_type(self, row) -> bool:
        return self._check_not_blank(row.get("PROCUREMENTTYPE"))

    # ─ Rule 8: ABCINDICATOR ─
    # NOTE: Validation intentionally skipped — column is fully blank in source data.
    # def validate_abc_indicator(self, row) -> bool:
    #     return self._check_not_blank(row.get("ABCINDICATOR"))

    # ─ Rule 9: IBPSTATUS ─
    def validate_ibp_status(self, row) -> bool:
        raw = row.get("IBPSTATUS")
        if self._is_blank(raw):
            return True
        return str(raw).strip().upper() == "IBP"

    # ─ Rule 10: XPLANTMATSTATUS ─
    def validate_xplant_mat_status(self, row) -> bool:
        raw = row.get("XPLANTMATSTATUS")
        if self._is_blank(raw):
            return True
        return str(raw).strip() == "2"

    def get_rules(self) -> dict:
        return {
            "MATERIALNUMBER":     self.validate_material_number,
            "PLANT":              self.validate_plant,
            "PRODUCTDESCRIPTION": self.validate_product_description,
            "PRODUCTTYPE":        self.validate_product_type,
            "PRODUCTHIERARCHY":   self.validate_product_hierarchy,
            "MRPTYPE":            self.validate_mrp_type,
            "PROCUREMENTTYPE":    self.validate_procurement_type,
            # ABCINDICATOR skipped — fully blank in source, no errors to mark
            "IBPSTATUS":          self.validate_ibp_status,
            "XPLANTMATSTATUS":    self.validate_xplant_mat_status,
        }


# ══════════════════════════════════════════════
#  Validator
# ══════════════════════════════════════════════
class PartTableValidator:
    """Reads the Excel file, applies rules, and produces an error map."""

    def __init__(self, filepath: str, valid_plants: list):
        self.filepath  = filepath
        self.engine    = RuleEngine(valid_plants)
        self.df        = pd.DataFrame()
        self.error_map = {}   # { row_idx: [col, col, ...] }

    def load(self):
        self.df = pd.read_excel(self.filepath, dtype=str)
        self.df.columns = [c.strip().upper() for c in self.df.columns]

    def validate(self):
        rules = self.engine.get_rules()
        for idx, row in self.df.iterrows():
            failed_cols = []
            for col, rule_fn in rules.items():
                if col not in self.df.columns:
                    continue
                try:
                    passed = rule_fn(row)
                except Exception:
                    passed = False
                if not passed:
                    failed_cols.append(col)
            if failed_cols:
                self.error_map[idx] = failed_cols

    def get_error_series(self) -> pd.Series:
        """
        Returns a Series of descriptive one-liner error messages per row.
        Each failed column maps to a human-readable message from ERROR_MESSAGES.
        Multiple failures are separated by ' | '.
        """
        result = {}
        for idx, cols in self.error_map.items():
            messages = [ERROR_MESSAGES.get(col, f"{col}: Validation failed") for col in cols]
            result[idx] = " | ".join(messages)
        return pd.Series(result, dtype=str)

    def get_errors_by_field(self) -> dict:
        """Returns { field_name: [row_idx, ...] } — rows that failed that specific field."""
        field_errors: dict = {}
        for row_idx, bad_cols in self.error_map.items():
            for col in bad_cols:
                field_errors.setdefault(col, []).append(row_idx)
        return field_errors


# ══════════════════════════════════════════════
#  Report Writer
# ══════════════════════════════════════════════
class ReportWriter:
    """Writes validated data to a multi-sheet Excel workbook."""

    SHEET_ALL     = "Full Data"
    SHEET_ERRORS  = "Error Rows"
    SHEET_SUMMARY = "Summary"
    SHEET_RULES   = "Rules"

    RULES_CONTENT = {
        "MATERIALNUMBER": [
            "Must not be blank.",
            "For FERT type: Material number must be in range 14000000000000 – 14999999999999.",
            "For HAWA type: Material number must be in range 15000000000000 – 15999999999999.",
        ],
        "PLANT": [
            "Must not be blank.",
            "Must be present in the Consolidated PL list (hardcoded in the script).",
        ],
        "PRODUCTDESCRIPTION": [
            "Must not be blank.",
        ],
        "PRODUCTTYPE": [
            "Must not be blank.",
            "Value must be either FERT or HAWA.",
        ],
        "PRODUCTHIERARCHY": [
            "Must not be blank.",
        ],
        "MRPTYPE": [
            "Must not be blank.",
            "Value must be either ND or PD.",
        ],
        "PROCUREMENTTYPE": [
            "Must not be blank.",
        ],
        "ABCINDICATOR": [
            "Must not be blank.",
            "NOTE: Validation currently skipped — column is fully blank in source data.",
        ],
        "IBPSTATUS": [
            "Allowed values: IBP or blank.",
            "Any other value is treated as an error.",
        ],
        "XPLANTMATSTATUS": [
            "Allowed values: 2 or blank.",
            "Any other value is treated as an error.",
        ],
    }

    def __init__(self, validator: PartTableValidator, output_path: str):
        self.validator   = validator
        self.output_path = output_path

    # ── helpers ──────────────────────────────
    def _write_header(self, ws, columns):
        for c_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=c_idx, value=col_name)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT

    def _write_rows(self, ws, df: pd.DataFrame):
        for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.font = BODY_FONT

    def _highlight(self, ws, df: pd.DataFrame, error_map: dict, col_index: dict):
        for df_idx, bad_cols in error_map.items():
            excel_row  = df_idx + 2
            total_cols = len(df.columns)
            for c in range(1, total_cols + 1):
                ws.cell(row=excel_row, column=c).fill = ROW_FILL
            for col_name in bad_cols:
                if col_name in col_index:
                    cell = ws.cell(row=excel_row, column=col_index[col_name])
                    cell.fill = RED_FILL
                    cell.font = ERR_FONT

    def _set_widths(self, ws):
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    # ── Summary sheet ─────────────────────────
    def _write_summary_sheet(self, wb, error_map: dict, total_rows: int):
        ws = wb.create_sheet(self.SHEET_SUMMARY)

        # ── Title: "Part Validation Summary" ──
        title_cell = ws.cell(row=1, column=1, value="Part Validation Summary")
        title_cell.font = Font(name="Arial", bold=True, size=14)
        title_cell.fill = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells("A1:E1")
        ws.row_dimensions[1].height = 24

        # ── Column headers ──
        HDR_SUMMARY_FILL = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
        headers = ["#", "Field Name", "Error Count", "% of Total Records"]
        col_widths = [6, 28, 16, 20]
        for c_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=c_idx, value=h)
            cell.fill = HDR_SUMMARY_FILL
            cell.font = Font(name="Arial", bold=True)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # ── Per-field rows ──
        col_error_counts = {}
        for bad_cols in error_map.values():
            for col in bad_cols:
                col_error_counts[col] = col_error_counts.get(col, 0) + 1

        row_num = 4
        for field_num, (col_name, count) in enumerate(col_error_counts.items(), start=1):
            pct = f"{(count / total_rows * 100):.2f}%" if total_rows > 0 else "0.00%"

            ws.cell(row=row_num, column=1, value=field_num).font = BODY_FONT
            ws.cell(row=row_num, column=2, value=col_name).font  = BODY_FONT
            ws.cell(row=row_num, column=3, value=count).font     = BODY_FONT
            ws.cell(row=row_num, column=4, value=pct).font       = BODY_FONT

            for c in range(1, 5):
                ws.cell(row=row_num, column=c).border    = THIN_BORDER
                ws.cell(row=row_num, column=c).alignment = Alignment(horizontal="center")

            row_num += 1

        # ── TOTAL row ──
        total_errors = sum(col_error_counts.values())
        total_pct    = f"{(total_errors / total_rows * 100):.2f}%" if total_rows > 0 else "0.00%"
        TOTAL_FILL   = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")

        ws.cell(row=row_num, column=2, value="TOTAL").font     = Font(name="Arial", bold=True)
        ws.cell(row=row_num, column=3, value=total_errors).font = Font(name="Arial", bold=True)
        ws.cell(row=row_num, column=4, value=total_pct).font   = Font(name="Arial", bold=True)
        for c in range(1, 5):
            ws.cell(row=row_num, column=c).fill      = TOTAL_FILL
            ws.cell(row=row_num, column=c).border    = THIN_BORDER
            ws.cell(row=row_num, column=c).alignment = Alignment(horizontal="center")

        # ── Spacer ──
        row_num += 2

        # ── Total Records / Records with Errors / Records Passing block ──
        records_with_errors = len(error_map)
        records_passing     = total_rows - records_with_errors

        stats = [
            ("Total Records:",        total_rows),
            ("Records with Errors:",  records_with_errors),
            ("Records Passing:",       records_passing),
        ]

        STATS_LABEL_FILL = PatternFill("solid", start_color="EDEDED", end_color="EDEDED")
        for label, value in stats:
            label_cell = ws.cell(row=row_num, column=1, value=label)
            label_cell.font      = Font(name="Arial", bold=True, size=10)
            label_cell.fill      = STATS_LABEL_FILL
            label_cell.border    = THIN_BORDER
            label_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(
                start_row=row_num, start_column=1,
                end_row=row_num,   end_column=2
            )

            value_cell = ws.cell(row=row_num, column=3, value=value)
            value_cell.font      = Font(name="Arial", size=10)
            value_cell.border    = THIN_BORDER
            value_cell.alignment = Alignment(horizontal="center", vertical="center")

            row_num += 1

        # ── Column widths ──
        for c_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = width

    # ── Per-field error sheets ────────────────
    def _write_field_error_sheets(self, wb, df: pd.DataFrame):
        """One sheet per errored field; shows ALL columns; row count matches Summary."""
        field_errors = self.validator.get_errors_by_field()

        for field_name, row_indices in field_errors.items():
            sheet_name = field_name[:31].replace("/", "-").replace("\\", "-").replace("*", "")
            ws = wb.create_sheet(sheet_name)

            subset = df.loc[row_indices].copy()
            self._write_header(ws, subset.columns)

            col_idx_map = {col: i for i, col in enumerate(subset.columns, start=1)}

            for excel_row, (orig_idx, row_data) in enumerate(subset.iterrows(), start=2):
                for c_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=excel_row, column=c_idx, value=value)
                    cell.font = BODY_FONT
                    cell.fill = ROW_FILL

                if field_name in col_idx_map:
                    bad_cell = ws.cell(row=excel_row, column=col_idx_map[field_name])
                    bad_cell.fill = RED_FILL
                    bad_cell.font = ERR_FONT

            self._set_widths(ws)

            note_row = len(subset) + 3
            ws.cell(
                row=note_row, column=1,
                value=f"Total error rows for '{field_name}': {len(subset)}",
            ).font = Font(name="Arial", italic=True, size=9, bold=True)

    # ── Rules sheet ───────────────────────────
    def _write_rules_sheet(self, wb):
        ws = wb.create_sheet(self.SHEET_RULES)

        title_cell = ws.cell(row=1, column=1, value="Part Table – Validation Rules")
        title_cell.font = Font(name="Arial", bold=True, size=13)
        title_cell.fill = TITLE_FILL
        ws.merge_cells("A1:C1")
        title_cell.alignment = Alignment(horizontal="center")

        for c_idx, h in enumerate(["#", "Field", "Rule Description"], start=1):
            cell = ws.cell(row=3, column=c_idx, value=h)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        current_row = 4
        rule_num    = 1

        for field, rules_list in self.RULES_CONTENT.items():
            num_rules = len(rules_list)

            for r_idx, rule_text in enumerate(rules_list):
                num_cell = ws.cell(row=current_row, column=1,
                                   value=rule_num if r_idx == 0 else "")
                num_cell.font   = Font(name="Arial", size=10, bold=(r_idx == 0))
                num_cell.fill   = RULE_FILL
                num_cell.border = THIN_BORDER
                num_cell.alignment = Alignment(horizontal="center", vertical="center")

                field_cell = ws.cell(row=current_row, column=2,
                                     value=field if r_idx == 0 else "")
                field_cell.font   = Font(name="Arial", size=10, bold=(r_idx == 0))
                field_cell.fill   = RULE_FILL
                field_cell.border = THIN_BORDER
                field_cell.alignment = Alignment(vertical="center")

                desc_cell = ws.cell(row=current_row, column=3, value=rule_text)
                desc_cell.font   = BODY_FONT
                desc_cell.border = THIN_BORDER
                desc_cell.alignment = Alignment(wrap_text=True, vertical="center")

                current_row += 1

            if num_rules > 1:
                merge_start = current_row - num_rules
                merge_end   = current_row - 1
                ws.merge_cells(
                    start_row=merge_start, start_column=1,
                    end_row=merge_end,     end_column=1,
                )
                ws.merge_cells(
                    start_row=merge_start, start_column=2,
                    end_row=merge_end,     end_column=2,
                )

            rule_num += 1

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 24
        ws.column_dimensions["C"].width = 70
        ws.row_dimensions[1].height = 22

    # ── Main write ────────────────────────────
    def write(self):
        v      = self.validator
        df     = v.df.copy()
        errors = v.error_map

        # ERROR_COLUMNS now shows descriptive one-liner messages
        df["ERROR_COLUMNS"] = df.index.map(lambda i: v.get_error_series().get(i, ""))
        col_index = {col: i for i, col in enumerate(df.columns, start=1)}

        # ── commented out: error_df used only for Error Rows sheet
        # error_df  = df[df.index.isin(errors.keys())].copy()

        wb = Workbook()

        # ── Sheet 1 – Full Data ──
        ws_all = wb.active
        ws_all.title = self.SHEET_ALL
        self._write_header(ws_all, df.columns)
        self._write_rows(ws_all, df)
        self._highlight(ws_all, df, errors, col_index)
        self._set_widths(ws_all)

        # ════════════════════════════════════════
        # Sheet 2 – Error Rows (COMMENTED OUT)
        # ════════════════════════════════════════
        # ws_err = wb.create_sheet(self.SHEET_ERRORS)
        # self._write_header(ws_err, error_df.columns)
        # self._write_rows(ws_err, error_df)
        # err_col_idx = {col: i for i, col in enumerate(error_df.columns, start=1)}
        # for sheet2_row, orig_idx in enumerate(error_df.index, start=2):
        #     bad_cols   = errors[orig_idx]
        #     total_cols = len(error_df.columns)
        #     for c in range(1, total_cols + 1):
        #         ws_err.cell(row=sheet2_row, column=c).fill = ROW_FILL
        #     for col_name in bad_cols:
        #         if col_name in err_col_idx:
        #             cell = ws_err.cell(row=sheet2_row, column=err_col_idx[col_name])
        #             cell.fill = RED_FILL
        #             cell.font = ERR_FONT
        # self._set_widths(ws_err)
        # ════════════════════════════════════════

        # ── Sheet 3 – Summary (updated format) ──
        self._write_summary_sheet(wb, errors, total_rows=len(df))

        # ── Sheets 4+ – One sheet per errored field ──
        self._write_field_error_sheets(wb, df)

        # ── Last sheet – Rules ──
        self._write_rules_sheet(wb)

        wb.save(self.output_path)
        print(f"\n✅  Output saved  → {self.output_path}")
        print(f"   Total rows    : {len(df)}")
        print(f"   Error rows    : {len(errors)}")
        print(f"   Field sheets  : {list(v.get_errors_by_field().keys())}")


# ══════════════════════════════════════════════
#  Orchestrator
# ══════════════════════════════════════════════
class PartTableProcessor:
    """Ties together loading, validation, and report writing."""

    def __init__(self, input_path: str, output_path: str, valid_plants: list):
        self.validator = PartTableValidator(input_path, valid_plants)
        self.writer    = ReportWriter(self.validator, output_path)

    def run(self):
        print("📂  Loading file …")
        self.validator.load()
        print(f"    Columns detected : {list(self.validator.df.columns)}")
        print("🔍  Validating rules …")
        self.validator.validate()
        print("📝  Writing report …")
        self.writer.write()


# ══════════════════════════════════════════════
#  Entry Point
# ══════════════════════════════════════════════
if __name__ == "__main__":
    processor = PartTableProcessor(
        input_path   = INPUT_FILE,
        output_path  = OUTPUT_FILE,
        valid_plants = VALID_PLANTS,
    )
    processor.run()

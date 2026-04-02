import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
#  FILE PATHS  –  update these
# ─────────────────────────────────────────────
SITE_INPUT_FILE  = r"D:/SEM-8/Data Rules Set Check/Data_rulesets_check/Excel_Files/Site.xlsx"
PART_INPUT_FILE  = r"D:/SEM-8/Data Rules Set Check/Data_rulesets_check/Excel_Files/Part.xlsx"
OUTPUT_FILE      = r"D:/SEM-8/Data Rules Set Check/Data_rulesets_check/Validated_Site.xlsx"


# ─────────────────────────────────────────────
#  CONSOLIDATED PL LIST
# ─────────────────────────────────────────────
VALID_PLANTS = [
    "1127", "1100", "1105", "1146", "1156", "1107", "1157", "1158", "1166", "1180",
    "1184", "1186", "1197", "1203", "1204", "1211", "1213", "1214", "1218", "1223",
    "1110", "1225", "1226", "1229", "1233", "1234", "1240", "1113", "1248", "1253",
    "1257", "1258", "1265", "1114", "1145", "1275", "1279", "1416", "1421", "1423",
    "1425", "1426", "1428", "1429", "1430", "1432", "1433", "1436", "1437", "1438",
    "1439", "1440", "1442", "1445", "1449", "1451", "1452", "1455", "1463", "1471",
    "1473", "1475", "1476", "1477", "1478", "1480", "1481", "1483", "1484", "1485",
    "1487", "1488", "1491", "1495", "1500", "1501", "1505", "1506", "1507", "1508",
    "1509", "1521", "1525", "1563", "1578", "1650", "1651", "1652", "1654", "1656",
    "1657", "1658", "1659", "1661", "1579", "1627", "1589", "1623", "1646", "1647",
    "1640", "1112", "5011637", "5123296", "5123742", "4007430", "1722", "1724", "1725",
    "1726", "1731", "1732", "1733", "1734", "1738", "1739", "1740", "1742", "1754",
    "1757", "1758", "1771", "1774", "1780", "1784", "1785", "1788", "1511", "1512",
    "1948", "1448", "4011702", "5013796", "5018849", "5011407", "5015073", "5011308",
    "5123531", "1642", "1638", "1104", "1104A", "1643", "1106", "1109", "1645", "1111",
    "1648", "1649", "1653", "1801", "1802", "2082", "2091", "2088", "2089", "2083",
    "2084", "2085", "2090", "2081", "2086", "2087", "1554", "1520", "1472", "1571",
    "1298", "1295", "1196", "1292", "1176", "1441", "1522", "1494", "1489", "1518",
    "1249", "1296", "1137", "1208", "1155", "1569", "1235", "1281", "1503", "1482",
    "1135", "1205", "1241", "1499", "1462", "1555", "1559", "1575", "2092", "1558",
    "1601", "1125", "1256", "1568",
]

# Valid values for COMPANYCODE
VALID_COMPANY_CODES = {"1001", "1006", "1009"}


# ─────────────────────────────────────────────
#  Colours
# ─────────────────────────────────────────────
RED_FILL    = PatternFill("solid", start_color="FF0000", end_color="FF0000")
ROW_FILL    = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
HDR_FILL    = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
RULE_FILL   = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
TITLE_FILL  = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
HDR_FONT    = Font(bold=True, name="Arial")
BODY_FONT   = Font(name="Arial", size=10)
ERR_FONT    = Font(name="Arial", size=10, bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


# ══════════════════════════════════════════════
#  Rule Engine  (returns failure reason strings)
# ══════════════════════════════════════════════
class SiteRuleEngine:
    """
    All column-level validation rules for the Site Table.
    Each validate_* method returns either:
      - None / empty string  → PASS
      - A short reason string → FAIL  (used verbatim in ERROR_COLUMNS)
    """

    def __init__(self, valid_plants: list, part_plants: set):
        self.valid_plants = set(str(p).strip() for p in valid_plants)
        self.part_plants  = set(str(p).strip() for p in part_plants)

    @staticmethod
    def _is_blank(value) -> bool:
        return pd.isna(value) or str(value).strip() == ""

    # ─ Rule 1: PLANT ─
    # Three sub-conditions checked in priority order:
    #   1a. Must not be blank
    #   1b. Must be present in Consolidated PL list
    #   1c. Must have an active part-site combination in Part master
    def validate_plant(self, row) -> str:
        val = str(row.get("PLANT", "")).strip()
        if not val or val == "nan":
            return "PLANT: Field is blank — plant code is mandatory"
        if val not in self.valid_plants:
            return f"PLANT: '{val}' is not present in the Consolidated PL list"
        if val not in self.part_plants:
            return f"PLANT: '{val}' has no active Part-Site combination in the Part master table"
        return ""   # PASS

    # ─ Rule 2: NAME ─
    def validate_name(self, row) -> str:
        if self._is_blank(row.get("NAME")):
            return "NAME: Field is blank — site name is mandatory"
        return ""

    # ─ Rule 3: ADDRESS ─
    def validate_address(self, row) -> str:
        if self._is_blank(row.get("ADDRESS")):
            return "ADDRESS: Field is blank — address is mandatory"
        return ""

    # ─ Rule 4: TCPL_PLANTTYPE ─
    def validate_tcpl_planttype(self, row) -> str:
        if self._is_blank(row.get("TCPL_PLANTTYPE")):
            return "TCPL_PLANTTYPE: Field is blank — plant type is mandatory"
        return ""

    # ─ Rule 5: COMPANYCODE ─
    def validate_companycode(self, row) -> str:
        val = row.get("COMPANYCODE", None)
        if self._is_blank(val):
            return "COMPANYCODE: Field is blank — company code is mandatory"
        if str(val).strip() not in VALID_COMPANY_CODES:
            return f"COMPANYCODE: '{str(val).strip()}' is invalid — must be one of 1001 / 1006 / 1009"
        return ""

    def get_rules(self) -> dict:
        """Returns { column_name: validate_fn } mapping."""
        return {
            "PLANT":          self.validate_plant,
            "NAME":           self.validate_name,
            "ADDRESS":        self.validate_address,
            "TCPL_PLANTTYPE": self.validate_tcpl_planttype,
            "COMPANYCODE":    self.validate_companycode,
        }


# ══════════════════════════════════════════════
#  Validator
# ══════════════════════════════════════════════
class SiteTableValidator:
    """Loads Site & Part Excel files, runs validation, builds error map."""

    def __init__(self, site_path: str, part_path: str, valid_plants: list):
        self.site_path    = site_path
        self.part_path    = part_path
        self.valid_plants = valid_plants
        self.df           = pd.DataFrame()
        self.part_plants  = set()
        self.error_map    = {}   # { row_idx: [col, col, ...] }  — col names only (for highlights)
        self.reason_map   = {}   # { row_idx: [reason_str, ...] } — full descriptions (for ERROR_COLUMNS)

    def load(self):
        self.df = pd.read_excel(self.site_path, dtype=str)
        self.df.columns = [c.strip().upper() for c in self.df.columns]

        part_df = pd.read_excel(self.part_path, dtype=str)
        part_df.columns = [c.strip().upper() for c in part_df.columns]

        if "PLANT" not in part_df.columns:
            raise ValueError("PLANT column not found in Part table.")

        self.part_plants = set(part_df["PLANT"].dropna().str.strip().tolist())
        print(f"    Part table plants loaded  : {len(self.part_plants)} unique values")

    def validate(self):
        engine = SiteRuleEngine(self.valid_plants, self.part_plants)
        rules  = engine.get_rules()

        for idx, row in self.df.iterrows():
            failed_cols   = []
            failed_reasons = []

            for col, rule_fn in rules.items():
                if col not in self.df.columns:
                    continue
                try:
                    reason = rule_fn(row)   # empty string = pass, non-empty = fail
                except Exception:
                    reason = f"{col}: Unexpected validation error"

                if reason:                  # non-empty → failure
                    failed_cols.append(col)
                    failed_reasons.append(reason)

            if failed_cols:
                self.error_map[idx]  = failed_cols
                self.reason_map[idx] = failed_reasons

    def get_error_series(self) -> pd.Series:
        """
        Returns a Series of pipe-separated, human-readable reason strings per row.
        Used to populate the ERROR_COLUMNS field in Full Data sheet.
        """
        result = {}
        for idx, reasons in self.reason_map.items():
            result[idx] = " | ".join(reasons)
        return pd.Series(result, dtype=str)

    def get_errors_by_field(self) -> dict:
        """Returns { field_name: [row_idx, ...] } — rows that failed that specific field."""
        field_errors: dict = {}
        for row_idx, bad_cols in self.error_map.items():
            for col in bad_cols:
                field_errors.setdefault(col, []).append(row_idx)
        return field_errors


# ══════════════════════════════════════════════
#  Report Writer
# ══════════════════════════════════════════════
class SiteReportWriter:
    """Builds the multi-sheet output Excel with colour highlights, summary,
       per-field error sheets, and a Rules sheet."""

    SHEET_ALL     = "Full Data"
    SHEET_ERRORS  = "Error Rows"
    SHEET_SUMMARY = "Summary"
    SHEET_RULES   = "Rules"

    RULES_CONTENT = {
        "PLANT": [
            "Must not be blank.",
            "Must be present in the Consolidated PL list (hardcoded in the script).",
            "Must have an active Part-Site combination in the Part master table (PLANT column).",
        ],
        "NAME": [
            "Must not be blank.",
        ],
        "ADDRESS": [
            "Must not be blank.",
        ],
        "TCPL_PLANTTYPE": [
            "Must not be blank.",
        ],
        "COMPANYCODE": [
            "Must not be blank.",
            "Value must be one of: 1001 / 1006 / 1009.",
        ],
    }

    def __init__(self, validator: SiteTableValidator, output_path: str):
        self.validator   = validator
        self.output_path = output_path

    # ── helpers ──────────────────────────────
    def _write_header(self, ws, columns):
        for c_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=c_idx, value=col_name)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT

    def _write_rows(self, ws, df: pd.DataFrame):
        for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.font = BODY_FONT

    def _highlight(self, ws, df: pd.DataFrame, error_map: dict, col_index: dict):
        for df_idx, bad_cols in error_map.items():
            excel_row  = df_idx + 2
            total_cols = len(df.columns)
            for c in range(1, total_cols + 1):
                ws.cell(row=excel_row, column=c).fill = ROW_FILL
            for col_name in bad_cols:
                if col_name in col_index:
                    cell = ws.cell(row=excel_row, column=col_index[col_name])
                    cell.fill = RED_FILL
                    cell.font = ERR_FONT

    def _set_widths(self, ws):
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    # ── Summary sheet (matches Part validation format) ────────────────────
    def _write_summary_sheet(self, wb, error_map: dict, total_rows: int):
        ws = wb.create_sheet(self.SHEET_SUMMARY)

        # ── Title ──
        title_cell = ws.cell(row=1, column=1, value="Site Validation Summary")
        title_cell.font = Font(name="Arial", bold=True, size=14)
        title_cell.fill = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells("A1:E1")
        ws.row_dimensions[1].height = 24

        # ── Column headers ──
        HDR_SUMMARY_FILL = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
        headers = ["#", "Field Name", "Error Count", "% of Total Records"]
        for c_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=c_idx, value=h)
            cell.fill = HDR_SUMMARY_FILL
            cell.font = Font(name="Arial", bold=True)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # ── Per-field rows ──
        col_error_counts = {}
        for bad_cols in error_map.values():
            for col in bad_cols:
                col_error_counts[col] = col_error_counts.get(col, 0) + 1

        row_num = 4
        for field_num, (col_name, count) in enumerate(col_error_counts.items(), start=1):
            pct = f"{(count / total_rows * 100):.2f}%" if total_rows > 0 else "0.00%"

            ws.cell(row=row_num, column=1, value=field_num).font = BODY_FONT
            ws.cell(row=row_num, column=2, value=col_name).font  = BODY_FONT
            ws.cell(row=row_num, column=3, value=count).font     = BODY_FONT
            ws.cell(row=row_num, column=4, value=pct).font       = BODY_FONT

            for c in range(1, 5):
                ws.cell(row=row_num, column=c).border    = THIN_BORDER
                ws.cell(row=row_num, column=c).alignment = Alignment(horizontal="center")

            row_num += 1

        # ── TOTAL row ──
        total_errors = sum(col_error_counts.values())
        total_pct    = f"{(total_errors / total_rows * 100):.2f}%" if total_rows > 0 else "0.00%"
        TOTAL_FILL   = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")

        ws.cell(row=row_num, column=2, value="TOTAL").font      = Font(name="Arial", bold=True)
        ws.cell(row=row_num, column=3, value=total_errors).font = Font(name="Arial", bold=True)
        ws.cell(row=row_num, column=4, value=total_pct).font    = Font(name="Arial", bold=True)
        for c in range(1, 5):
            ws.cell(row=row_num, column=c).fill      = TOTAL_FILL
            ws.cell(row=row_num, column=c).border    = THIN_BORDER
            ws.cell(row=row_num, column=c).alignment = Alignment(horizontal="center")

        # ── Spacer then stats block ──
        row_num += 2

        records_with_errors = len(error_map)
        records_passing     = total_rows - records_with_errors

        stats = [
            ("Total Records:",       total_rows),
            ("Records with Errors:", records_with_errors),
            ("Records Passing:",     records_passing),
        ]

        STATS_LABEL_FILL = PatternFill("solid", start_color="EDEDED", end_color="EDEDED")
        for label, value in stats:
            label_cell = ws.cell(row=row_num, column=1, value=label)
            label_cell.font      = Font(name="Arial", bold=True, size=10)
            label_cell.fill      = STATS_LABEL_FILL
            label_cell.border    = THIN_BORDER
            label_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(
                start_row=row_num, start_column=1,
                end_row=row_num,   end_column=2
            )

            value_cell = ws.cell(row=row_num, column=3, value=value)
            value_cell.font      = Font(name="Arial", size=10)
            value_cell.border    = THIN_BORDER
            value_cell.alignment = Alignment(horizontal="center", vertical="center")

            row_num += 1

        # ── Column widths ──
        col_widths = [6, 28, 16, 20]
        for c_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = width

    # ── Per-field error sheets ────────────────
    def _write_field_error_sheets(self, wb, df: pd.DataFrame):
        """One sheet per errored field; whole row light yellow, failing cell red."""
        field_errors = self.validator.get_errors_by_field()

        for field_name, row_indices in field_errors.items():
            sheet_name = field_name[:31].replace("/", "-").replace("\\", "-").replace("*", "")
            ws = wb.create_sheet(sheet_name)

            subset = df.loc[row_indices].copy()
            self._write_header(ws, subset.columns)

            col_idx_map = {col: i for i, col in enumerate(subset.columns, start=1)}

            for excel_row, (orig_idx, row_data) in enumerate(subset.iterrows(), start=2):
                for c_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=excel_row, column=c_idx, value=value)
                    cell.font = BODY_FONT
                    cell.fill = ROW_FILL

                if field_name in col_idx_map:
                    bad_cell = ws.cell(row=excel_row, column=col_idx_map[field_name])
                    bad_cell.fill = RED_FILL
                    bad_cell.font = ERR_FONT

            self._set_widths(ws)

            note_row = len(subset) + 3
            ws.cell(
                row=note_row, column=1,
                value=f"Total error rows for '{field_name}': {len(subset)}",
            ).font = Font(name="Arial", italic=True, size=9, bold=True)

    # ── Rules sheet ───────────────────────────
    def _write_rules_sheet(self, wb):
        ws = wb.create_sheet(self.SHEET_RULES)

        title_cell = ws.cell(row=1, column=1, value="Site Table – Validation Rules")
        title_cell.font = Font(name="Arial", bold=True, size=13)
        title_cell.fill = TITLE_FILL
        ws.merge_cells("A1:C1")
        title_cell.alignment = Alignment(horizontal="center")

        for c_idx, h in enumerate(["#", "Field", "Rule Description"], start=1):
            cell = ws.cell(row=3, column=c_idx, value=h)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        current_row = 4
        rule_num    = 1

        for field, rules_list in self.RULES_CONTENT.items():
            num_rules = len(rules_list)

            for r_idx, rule_text in enumerate(rules_list):
                num_cell = ws.cell(row=current_row, column=1,
                                   value=rule_num if r_idx == 0 else "")
                num_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                num_cell.fill      = RULE_FILL
                num_cell.border    = THIN_BORDER
                num_cell.alignment = Alignment(horizontal="center", vertical="center")

                field_cell = ws.cell(row=current_row, column=2,
                                     value=field if r_idx == 0 else "")
                field_cell.font      = Font(name="Arial", size=10, bold=(r_idx == 0))
                field_cell.fill      = RULE_FILL
                field_cell.border    = THIN_BORDER
                field_cell.alignment = Alignment(vertical="center")

                desc_cell = ws.cell(row=current_row, column=3, value=rule_text)
                desc_cell.font      = BODY_FONT
                desc_cell.border    = THIN_BORDER
                desc_cell.alignment = Alignment(wrap_text=True, vertical="center")

                current_row += 1

            if num_rules > 1:
                merge_start = current_row - num_rules
                merge_end   = current_row - 1
                ws.merge_cells(start_row=merge_start, start_column=1,
                               end_row=merge_end,     end_column=1)
                ws.merge_cells(start_row=merge_start, start_column=2,
                               end_row=merge_end,     end_column=2)

            rule_num += 1

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 70
        ws.row_dimensions[1].height = 22

    # ── Main write ────────────────────────────
    def write(self):
        v  = self.validator
        df = v.df.copy()

        # ERROR_COLUMNS: pipe-separated, condition-specific failure reasons
        error_series        = v.get_error_series()
        df["ERROR_COLUMNS"] = df.index.map(
            lambda i: error_series.get(i, "") if i in error_series.index else ""
        )

        col_index = {col: i for i, col in enumerate(df.columns, start=1)}

        # ── commented out: error_df used only for Error Rows sheet
        # error_df  = df[df.index.isin(v.error_map.keys())].copy()

        wb = Workbook()

        # ── Sheet 1 – Full Data ──
        ws_all = wb.active
        ws_all.title = self.SHEET_ALL
        self._write_header(ws_all, df.columns)
        self._write_rows(ws_all, df)
        self._highlight(ws_all, df, v.error_map, col_index)
        self._set_widths(ws_all)

        # ════════════════════════════════════════
        # Sheet 2 – Error Rows (COMMENTED OUT)
        # ════════════════════════════════════════
        # error_df  = df[df.index.isin(v.error_map.keys())].copy()
        # ws_err = wb.create_sheet(self.SHEET_ERRORS)
        # self._write_header(ws_err, error_df.columns)
        # self._write_rows(ws_err, error_df)
        # err_col_idx = {col: i for i, col in enumerate(error_df.columns, start=1)}
        # for sheet2_row, orig_idx in enumerate(error_df.index, start=2):
        #     bad_cols   = v.error_map[orig_idx]
        #     total_cols = len(error_df.columns)
        #     for c in range(1, total_cols + 1):
        #         ws_err.cell(row=sheet2_row, column=c).fill = ROW_FILL
        #     for col_name in bad_cols:
        #         if col_name in err_col_idx:
        #             cell = ws_err.cell(row=sheet2_row, column=err_col_idx[col_name])
        #             cell.fill = RED_FILL
        #             cell.font = ERR_FONT
        # self._set_widths(ws_err)
        # ════════════════════════════════════════

        # ── Sheet 3 – Summary (updated format) ──
        self._write_summary_sheet(wb, v.error_map, total_rows=len(df))

        # ── Sheets 4+ – One sheet per errored field ──
        self._write_field_error_sheets(wb, df)

        # ── Last sheet – Rules ──
        self._write_rules_sheet(wb)

        wb.save(self.output_path)
        print(f"\n✅  Output saved  → {self.output_path}")
        print(f"   Total rows    : {len(df)}")
        print(f"   Error rows    : {len(v.error_map)}")
        print(f"   Field sheets  : {list(v.get_errors_by_field().keys())}")


# ══════════════════════════════════════════════
#  Orchestrator
# ══════════════════════════════════════════════
class SiteTableProcessor:
    """Ties together loading, validation, and report writing."""

    def __init__(self, site_path: str, part_path: str, output_path: str, valid_plants: list):
        self.validator = SiteTableValidator(site_path, part_path, valid_plants)
        self.writer    = SiteReportWriter(self.validator, output_path)

    def run(self):
        print("📂  Loading files …")
        self.validator.load()
        print(f"    Site columns detected : {list(self.validator.df.columns)}")
        print("🔍  Validating rules …")
        self.validator.validate()
        print("📝  Writing report …")
        self.writer.write()


# ══════════════════════════════════════════════
#  Entry Point
# ══════════════════════════════════════════════
if __name__ == "__main__":
    processor = SiteTableProcessor(
        site_path    = SITE_INPUT_FILE,
        part_path    = PART_INPUT_FILE,
        output_path  = OUTPUT_FILE,
        valid_plants = VALID_PLANTS,
    )
    processor.run()

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
#  FILE PATHS  –  update these
# ─────────────────────────────────────────────
CUSTOMER_INPUT_FILE = r"C:\Users\M sD\Downloads\Data_rulesets_check\Excel_Files\Customer.xlsx"
SITE_INPUT_FILE     = r"C:\Users\M sD\Downloads\Data_rulesets_check\Excel_Files\Site.xlsx"
OUTPUT_FILE         = r"C:\Users\M sD\Downloads\Data_rulesets_check\Output_Files\Validated_Customer.xlsx"


# ─────────────────────────────────────────────
#  Colours
# ─────────────────────────────────────────────
RED_FILL  = PatternFill("solid", start_color="FF0000", end_color="FF0000")
ROW_FILL  = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
HDR_FILL  = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
HDR_FONT  = Font(bold=True, name="Arial")
BODY_FONT = Font(name="Arial", size=10)
ERR_FONT  = Font(name="Arial", size=10, bold=True, color="FFFFFF")


# ══════════════════════════════════════════════
#  Rule Engine
# ══════════════════════════════════════════════
class CustomerRuleEngine:
    """All column-level validation rules for the Customer Table."""

    def __init__(self, site_plants: set):
        self.site_plants = set(str(p).strip() for p in site_plants)

    @staticmethod
    def _is_blank(value) -> bool:
        return pd.isna(value) or str(value).strip() == ""

    def _check_not_blank(self, value) -> bool:
        return not self._is_blank(value)

    # ─ Rule 1: CUSTOMER ─
    def validate_customer(self, row) -> bool:
        return self._check_not_blank(row.get("CUSTOMER", None))

    # ─ Rule 2: CUSTOMERNAME ─
    def validate_customername(self, row) -> bool:
        return self._check_not_blank(row.get("CUSTOMERNAME", None))

    # ─ Rule 3: SUPPLYINGPLANT ─
    def validate_supplyingplant(self, row) -> bool:
        val = row.get("SUPPLYINGPLANT", None)
        if self._is_blank(val):
            return False
        return str(val).strip() in self.site_plants

    # ─ Rule 4: CUSTOMERGROUP ─
    def validate_customergroup(self, row) -> bool:
        return self._check_not_blank(row.get("CUSTOMERGROUP", None))

    # ─ Rule 5: CUSTOMERGROUP_NAME ─
    def validate_customergroup_name(self, row) -> bool:
        return self._check_not_blank(row.get("CUSTOMERGROUP_NAME", None))

    # ─ Rule 6: SALESORGANIZATION ─
    def validate_salesorganization(self, row) -> bool:
        return self._check_not_blank(row.get("SALESORGANIZATION", None))

    # ─ Rule 7: CUSTOMERGROUP1 ─
    def validate_customergroup1(self, row) -> bool:
        return self._check_not_blank(row.get("CUSTOMERGROUP1", None))

    # ─ Rule 8: CUSTOMERGROUP1_NAME ─
    def validate_customergroup1_name(self, row) -> bool:
        return self._check_not_blank(row.get("CUSTOMERGROUP1_NAME", None))

    # ─ Rule 9: COUNTRY_CODE ─
    def validate_country_code(self, row) -> bool:
        return self._check_not_blank(row.get("COUNTRY_CODE", None))

    # ─ Rule 10: COUNTRY_NAME ─
    def validate_country_name(self, row) -> bool:
        return self._check_not_blank(row.get("COUNTRY_NAME", None))

    # ─ Rule 11: CHANNEL_CODE ─
    def validate_channel_code(self, row) -> bool:
        return self._check_not_blank(row.get("CHANNEL_CODE", None))

    # ─ Rule 12: GLOBAL_CHANNEL_NAME ─
    def validate_global_channel_name(self, row) -> bool:
        return self._check_not_blank(row.get("GLOBAL_CHANNEL_NAME", None))

    # ─ Rule 13: CHANNEL ─
    def validate_channel(self, row) -> bool:
        return self._check_not_blank(row.get("CHANNEL", None))

    # ─ Rule 14: CHANNEL_DESCRIPTION ─
    def validate_channel_description(self, row) -> bool:
        return self._check_not_blank(row.get("CHANNEL_DESCRIPTION", None))

    # ─ Rule 15: DIVISION ─
    def validate_division(self, row) -> bool:
        return self._check_not_blank(row.get("DIVISION", None))

    # ─ Rule 16: DIVISIONDESCRIPTION ─
    def validate_divisiondescription(self, row) -> bool:
        return self._check_not_blank(row.get("DIVISIONDESCRIPTION", None))

    # ─ Rule 17: SUB_CHANNEL_CODE_JDA_REPORTING ─
    def validate_sub_channel_code_jda_reporting(self, row) -> bool:
        return self._check_not_blank(row.get("SUB_CHANNEL_CODE_JDA_REPORTING", None))

    # ─ Rule 18: SUB_CHANNEL_DESC_JDA_REPORTING ─
    def validate_sub_channel_desc_jda_reporting(self, row) -> bool:
        return self._check_not_blank(row.get("SUB_CHANNEL_DESC_JDA_REPORTING", None))

    # ─ Rule 19: REGION_CODE ─
    def validate_region_code(self, row) -> bool:
        return self._check_not_blank(row.get("REGION_CODE", None))

    # ─ Rule 20: REGION_NAME ─
    def validate_region_name(self, row) -> bool:
        return self._check_not_blank(row.get("REGION_NAME", None))

    # ─ Rule 21: ASM_CODE ─
    def validate_asm_code(self, row) -> bool:
        return self._check_not_blank(row.get("ASM_CODE", None))

    # ─ Rule 22: AREA_NAME ─
    def validate_area_name(self, row) -> bool:
        return self._check_not_blank(row.get("AREA_NAME", None))

    # ─ Rule 23: MARKET_CODE ─
    def validate_market_code(self, row) -> bool:
        return self._check_not_blank(row.get("MARKET_CODE", None))

    # ─ Rule 24: MARKET_NAME ─
    def validate_market_name(self, row) -> bool:
        return self._check_not_blank(row.get("MARKET_NAME", None))

    # ─ Rule 25: CLUSTER_MANAGER_CODE ─
    def validate_cluster_manager_code(self, row) -> bool:
        return self._check_not_blank(row.get("CLUSTER_MANAGER_CODE", None))

    # ─ Rule 26: CM_NAME ─
    def validate_cm_name(self, row) -> bool:
        return self._check_not_blank(row.get("CM_NAME", None))

    # ─ Rule 27: SALESHIERARCHY ─
    def validate_saleshierarchy(self, row) -> bool:
        return self._check_not_blank(row.get("SALESHIERARCHY", None))

    def get_rules(self) -> dict:
        return {
            "CUSTOMER":                           self.validate_customer,
            "CUSTOMERNAME":                       self.validate_customername,
            "SUPPLYINGPLANT":                     self.validate_supplyingplant,
            "CUSTOMERGROUP":                      self.validate_customergroup,
            "CUSTOMERGROUP_NAME":                 self.validate_customergroup_name,
            "SALESORGANIZATION":                  self.validate_salesorganization,
            "CUSTOMERGROUP1":                     self.validate_customergroup1,
            "CUSTOMERGROUP1_NAME":                self.validate_customergroup1_name,
            "COUNTRY_CODE":                       self.validate_country_code,
            "COUNTRY_NAME":                       self.validate_country_name,
            "CHANNEL_CODE":                       self.validate_channel_code,
            "GLOBAL_CHANNEL_NAME":                self.validate_global_channel_name,
            "CHANNEL":                            self.validate_channel,
            "CHANNEL_DESCRIPTION":                self.validate_channel_description,
            "DIVISION":                           self.validate_division,
            "DIVISIONDESCRIPTION":                self.validate_divisiondescription,
            "SUB_CHANNEL_CODE_JDA_REPORTING":     self.validate_sub_channel_code_jda_reporting,
            "SUB_CHANNEL_DESC_JDA_REPORTING":     self.validate_sub_channel_desc_jda_reporting,
            "REGION_CODE":                        self.validate_region_code,
            "REGION_NAME":                        self.validate_region_name,
            "ASM_CODE":                           self.validate_asm_code,
            "AREA_NAME":                          self.validate_area_name,
            "MARKET_CODE":                        self.validate_market_code,
            "MARKET_NAME":                        self.validate_market_name,
            "CLUSTER_MANAGER_CODE":               self.validate_cluster_manager_code,
            "CM_NAME":                            self.validate_cm_name,
            "SALESHIERARCHY":                     self.validate_saleshierarchy,
        }


# ══════════════════════════════════════════════
#  Validator
# ══════════════════════════════════════════════
class CustomerTableValidator:
    """Loads Customer & Site Excel files, runs validation, builds error map."""

    def __init__(self, customer_path: str, site_path: str):
        self.customer_path = customer_path
        self.site_path     = site_path
        self.df            = pd.DataFrame()
        self.site_plants   = set()
        self.error_map     = {}

    def load(self):
        self.df = pd.read_excel(self.customer_path, dtype=str)
        self.df.columns = [c.strip().upper() for c in self.df.columns]

        site_df = pd.read_excel(self.site_path, dtype=str)
        site_df.columns = [c.strip().upper() for c in site_df.columns]

        if "PLANT" not in site_df.columns:
            raise ValueError("PLANT column not found in Site table.")

        self.site_plants = set(site_df["PLANT"].dropna().str.strip().tolist())
        print(f"    Site table plants loaded  : {len(self.site_plants)} unique values")

    def validate(self):
        engine = CustomerRuleEngine(self.site_plants)
        rules  = engine.get_rules()
        for idx, row in self.df.iterrows():
            failed = []
            for col, rule_fn in rules.items():
                if col not in self.df.columns:
                    continue
                try:
                    passed = rule_fn(row)
                except Exception:
                    passed = False
                if not passed:
                    failed.append(col)
            if failed:
                self.error_map[idx] = failed

    def get_error_series(self) -> pd.Series:
        return pd.Series(
            {idx: ", ".join(cols) for idx, cols in self.error_map.items()},
            dtype=str
        )


# ══════════════════════════════════════════════
#  Report Writer
# ══════════════════════════════════════════════
class CustomerReportWriter:
    """Builds the 3-sheet output Excel with colour highlights and summary."""

    SHEET_ALL     = "Full Data"
    SHEET_ERRORS  = "Error Rows"
    SHEET_SUMMARY = "Summary"

    def __init__(self, validator: CustomerTableValidator, output_path: str):
        self.validator   = validator
        self.output_path = output_path

    def _write_header(self, ws, columns):
        for c_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=c_idx, value=col_name)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT

    def _write_rows(self, ws, df: pd.DataFrame):
        for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.font = BODY_FONT

    def _highlight(self, ws, df: pd.DataFrame, error_map: dict, col_index: dict):
        for df_idx, bad_cols in error_map.items():
            excel_row  = df_idx + 2
            total_cols = len(df.columns)
            for c in range(1, total_cols + 1):
                ws.cell(row=excel_row, column=c).fill = ROW_FILL
            for col_name in bad_cols:
                if col_name in col_index:
                    cell = ws.cell(row=excel_row, column=col_index[col_name])
                    cell.fill = RED_FILL
                    cell.font = ERR_FONT

    def _set_widths(self, ws):
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    def _write_summary_sheet(self, wb, error_map: dict):
        ws = wb.create_sheet(self.SHEET_SUMMARY)

        # Title
        ws.cell(row=1, column=1, value="Summary Output").font = Font(name="Arial", bold=True, size=13)
        ws.cell(row=1, column=2, value="In one sheet - Summary").font = Font(name="Arial", size=10)

        # Header
        hdr_fill = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
        for c_idx, h in enumerate(["Extract", "Field", "Count of Errors"], start=1):
            cell = ws.cell(row=3, column=c_idx, value=h)
            cell.fill = hdr_fill
            cell.font = Font(name="Arial", bold=True)

        # Count errors per column
        col_error_counts = {}
        for bad_cols in error_map.values():
            for col in bad_cols:
                col_error_counts[col] = col_error_counts.get(col, 0) + 1

        # Data rows
        row_num = 4
        total   = 0
        for col_name, count in col_error_counts.items():
            ws.cell(row=row_num, column=1, value="Customer").font = BODY_FONT
            ws.cell(row=row_num, column=2, value=col_name).font = BODY_FONT
            ws.cell(row=row_num, column=3, value=count).font = BODY_FONT
            total   += count
            row_num += 1

        # Total row
        total_fill = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
        ws.cell(row=row_num, column=2, value="Total").font = Font(name="Arial", bold=True)
        ws.cell(row=row_num, column=3, value=total).font = Font(name="Arial", bold=True)
        for c in range(1, 4):
            ws.cell(row=row_num, column=c).fill = total_fill

        # Note
        ws.cell(
            row=row_num + 2, column=1,
            value="Each field will have a separate sheet displaying all rows with errors"
        ).font = Font(name="Arial", italic=True, size=9)

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 18

    def write(self):
        v  = self.validator
        df = v.df.copy()
        error_series = v.get_error_series()
        df["ERROR_COLUMNS"] = df.index.map(lambda i: error_series.get(i, "") if i in error_series.index else "")

        col_index = {col: i for i, col in enumerate(df.columns, start=1)}
        error_df  = df[df.index.isin(v.error_map.keys())].copy()

        wb = Workbook()

        # ── Sheet 1: Full Data ────────────────
        ws_all = wb.active
        ws_all.title = self.SHEET_ALL
        self._write_header(ws_all, df.columns)
        self._write_rows(ws_all, df)
        self._highlight(ws_all, df, v.error_map, col_index)
        self._set_widths(ws_all)

        # ── Sheet 2: Error Rows ───────────────
        ws_err = wb.create_sheet(self.SHEET_ERRORS)
        self._write_header(ws_err, error_df.columns)
        self._write_rows(ws_err, error_df)
        err_col_idx = {col: i for i, col in enumerate(error_df.columns, start=1)}
        for sheet2_row, orig_idx in enumerate(error_df.index, start=2):
            bad_cols   = v.error_map[orig_idx]
            total_cols = len(error_df.columns)
            for c in range(1, total_cols + 1):
                ws_err.cell(row=sheet2_row, column=c).fill = ROW_FILL
            for col_name in bad_cols:
                if col_name in err_col_idx:
                    cell = ws_err.cell(row=sheet2_row, column=err_col_idx[col_name])
                    cell.fill = RED_FILL
                    cell.font = ERR_FONT
        self._set_widths(ws_err)

        # ── Sheet 3: Summary ──────────────────
        self._write_summary_sheet(wb, v.error_map)

        wb.save(self.output_path)
        print(f"\n✅  Output saved  → {self.output_path}")
        print(f"   Total rows    : {len(df)}")
        print(f"   Error rows    : {len(error_df)}")


# ══════════════════════════════════════════════
#  Orchestrator
# ══════════════════════════════════════════════
class CustomerTableProcessor:
    """Ties together loading, validation, and report writing."""

    def __init__(self, customer_path: str, site_path: str, output_path: str):
        self.validator = CustomerTableValidator(customer_path, site_path)
        self.writer    = CustomerReportWriter(self.validator, output_path)

    def run(self):
        print("📂  Loading files …")
        self.validator.load()
        print(f"    Customer columns detected : {list(self.validator.df.columns)}")
        print("🔍  Validating rules …")
        self.validator.validate()
        print("📝  Writing report …")
        self.writer.write()


# ══════════════════════════════════════════════
#  Entry Point
# ══════════════════════════════════════════════
if __name__ == "__main__":
    processor = CustomerTableProcessor(
        customer_path = CUSTOMER_INPUT_FILE,
        site_path     = SITE_INPUT_FILE,
        output_path   = OUTPUT_FILE,
    )
    processor.run()
